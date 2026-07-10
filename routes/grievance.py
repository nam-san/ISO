import io
from datetime import datetime, date

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, abort
from flask_login import login_required, current_user

from models import db, Grievance, Department, User, AuditTrail

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

grievance_bp = Blueprint('grievance', __name__)

CATEGORIES = ['임금', '근로시간', '직장내괴롭힘', '성희롱', '안전보건', '복리후생', '인사', '기타']
SEVERITY_KR = {'high': '높음', 'normal': '보통', 'low': '낮음'}
STATUS_KR = {'open': '접수', 'investigating': '조사중', 'resolved': '처리완료', 'closed': '종결'}


def _can_handle():
    """고충 처리(열람·조치) 권한: 매니저/관리자만 (비밀유지)."""
    return current_user.is_manager()


def _next_number():
    year = datetime.now().year
    last = Grievance.query.filter(
        Grievance.grievance_number.like(f'GRV-{year}-%')
    ).order_by(Grievance.id.desc()).first()
    seq = int(last.grievance_number.split('-')[-1]) + 1 if last else 1
    return f'GRV-{year}-{seq:04d}'


# ── 목록 ─────────────────────────────────────────────────────
@grievance_bp.route('/')
@login_required
def index():
    status = request.args.get('status', '')

    if _can_handle():
        # 처리권자: 전체 열람
        query = Grievance.query
        if status:
            query = query.filter_by(status=status)
        grievances = query.order_by(Grievance.receipt_date.desc()).all()
        mine = False
    else:
        # 일반 직원: 본인이 실명으로 제기한 건만 (익명건은 본인도 추적 불가)
        grievances = Grievance.query.filter_by(
            reporter_id=current_user.id, is_anonymous=False
        ).order_by(Grievance.receipt_date.desc()).all()
        mine = True

    today = date.today()
    stats = {
        'open':          Grievance.query.filter_by(status='open').count(),
        'investigating': Grievance.query.filter_by(status='investigating').count(),
        'resolved':      Grievance.query.filter_by(status='resolved').count(),
        'closed':        Grievance.query.filter_by(status='closed').count(),
    } if _can_handle() else {}

    return render_template('grievance/index.html',
        grievances=grievances, mine=mine, can_handle=_can_handle(),
        status=status, stats=stats, today=today,
        severity_kr=SEVERITY_KR, status_kr=STATUS_KR)


# ── 고충 제기 (모든 로그인 사용자) ───────────────────────────
@grievance_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if request.method == 'POST':
        is_anon = request.form.get('is_anonymous') == 'on'
        receipt = request.form.get('receipt_date')
        receipt_date = datetime.strptime(receipt, '%Y-%m-%d').date() if receipt else date.today()

        g = Grievance(
            grievance_number = _next_number(),
            receipt_date     = receipt_date,
            category         = request.form.get('category', '기타'),
            title            = request.form['title'],
            content          = request.form['content'],
            is_anonymous     = is_anon,
            reporter_id      = None if is_anon else current_user.id,
            reporter_name    = None if is_anon else current_user.name,
            department_id    = None if is_anon else (current_user.department_id),
            contact_info     = request.form.get('contact_info') if not is_anon or request.form.get('contact_info') else None,
            severity         = request.form.get('severity', 'normal'),
            status           = 'open',
        )
        db.session.add(g)
        db.session.flush()

        # 감사로그: 익명 보호를 위해 사용자 ID 미기록
        log = AuditTrail(
            user_id = None if is_anon else current_user.id,
            action  = '고충접수',
            target_type = 'grievance', target_id = g.id,
            target_name = f'{g.grievance_number} ({g.category})'
        )
        db.session.add(log)
        db.session.commit()

        if is_anon:
            flash(f'고충이 익명으로 접수되었습니다. (접수번호 {g.grievance_number}) '
                  f'익명 접수 건은 추적되지 않으니 번호를 보관하세요.', 'success')
        else:
            flash(f'고충이 접수되었습니다. (접수번호 {g.grievance_number})', 'success')
        return redirect(url_for('grievance.index'))

    today = date.today()
    return render_template('grievance/new.html',
        today=today, categories=CATEGORIES)


# ── 상세 / 처리 ──────────────────────────────────────────────
@grievance_bp.route('/<int:grievance_id>', methods=['GET', 'POST'])
@login_required
def detail(grievance_id):
    g = Grievance.query.get_or_404(grievance_id)

    # 접근 제어: 처리권자 OR (실명 제기한 본인)
    is_owner = (not g.is_anonymous and g.reporter_id == current_user.id)
    if not (_can_handle() or is_owner):
        abort(403)

    if request.method == 'POST':
        if not _can_handle():
            flash('처리 권한이 없습니다.', 'danger')
            return redirect(url_for('grievance.detail', grievance_id=g.id))

        g.status        = request.form.get('status', g.status)
        g.handler_id    = request.form.get('handler_id', type=int) or g.handler_id
        g.investigation = request.form.get('investigation')
        g.action_taken  = request.form.get('action_taken')
        g.result        = request.form.get('result')
        due = request.form.get('due_date')
        g.due_date = datetime.strptime(due, '%Y-%m-%d').date() if due else None
        if g.status in ('resolved', 'closed') and not g.resolved_date:
            g.resolved_date = date.today()
        sat = request.form.get('satisfaction', type=int)
        if sat:
            g.satisfaction = sat

        db.session.add(AuditTrail(user_id=current_user.id, action='고충처리',
                                  target_type='grievance', target_id=g.id,
                                  target_name=g.grievance_number))
        db.session.commit()
        flash('고충 처리 내용이 저장되었습니다.', 'success')
        return redirect(url_for('grievance.detail', grievance_id=g.id))

    handlers = User.query.filter(User.role.in_(['admin', 'manager'])).order_by(User.name).all()
    today = date.today()
    return render_template('grievance/detail.html',
        g=g, is_owner=is_owner, can_handle=_can_handle(),
        handlers=handlers, categories=CATEGORIES, today=today,
        severity_kr=SEVERITY_KR, status_kr=STATUS_KR)


# ── Excel 출력 (처리권자) ────────────────────────────────────
@grievance_bp.route('/export')
@login_required
def export_excel():
    if not _can_handle():
        abort(403)
    grievances = Grievance.query.order_by(Grievance.receipt_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '근로자 고충처리 대장'

    hdr_fill = PatternFill('solid', fgColor='1a3a5c')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:I1')
    ws['A1'] = '근로자 고충처리 대장 (비밀유지)'
    ws['A1'].font = Font(bold=True, size=14, color='1a3a5c')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['접수번호', '접수일', '분류', '제목', '제기자', '심각도', '상태', '처리담당', '처리완료일']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin

    for i, g in enumerate(grievances, 3):
        reporter = '익명' if g.is_anonymous else (g.reporter_name or '-')
        row = [g.grievance_number, str(g.receipt_date), g.category, g.title, reporter,
               SEVERITY_KR.get(g.severity, g.severity), STATUS_KR.get(g.status, g.status),
               g.handler.name if g.handler else '-',
               str(g.resolved_date) if g.resolved_date else '-']
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = left if col == 4 else center
            c.border = thin

    widths = [16, 12, 14, 30, 12, 9, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(2, i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output,
        download_name=f'근로자고충처리대장_{datetime.now().strftime("%Y%m%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── 삭제 (비밀유지·기록보존 원칙상 관리자 전용) ──────────────
@grievance_bp.route('/<int:grievance_id>/delete', methods=['POST'])
@login_required
def delete(grievance_id):
    g = Grievance.query.get_or_404(grievance_id)
    if not current_user.is_admin():
        flash('고충처리 기록은 비밀유지·보존 원칙에 따라 관리자만 삭제할 수 있습니다.', 'danger')
        return redirect(url_for('grievance.detail', grievance_id=grievance_id))
    db.session.add(AuditTrail(user_id=current_user.id, action='고충삭제',
        target_type='grievance', target_id=g.id, target_name=g.grievance_number))
    db.session.delete(g)
    db.session.commit()
    flash('고충처리 기록이 삭제되었습니다.', 'info')
    return redirect(url_for('grievance.index'))
