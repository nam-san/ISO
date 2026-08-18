from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, DesignChange, ChangeReview, ChangeAttachment, Department, User, AuditTrail
from datetime import datetime, date
import os, io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

design_bp = Blueprint('design', __name__)

IMAGE_EXTS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# 관련 서류 첨부 슬롯 — 표준문서 제/개정 5종 + 시험결과 4종 + 기타
DOC_SLOTS = [
    ('doc_inspection', '검사기준서'), ('doc_process', '공정도'),
    ('doc_drawing', '도면/시방서'), ('doc_control', '관리계획서'),
    ('doc_workstd', '작업표준서'),
]
TEST_SLOTS = [
    ('test_report', '검사성적서'), ('test_reliability', '신뢰성시험결과서'),
    ('test_ecn', 'ECN'), ('test_risk', '리스크파악'),
]
ALL_SLOTS = ['before', 'after', 'etc'] + [k for k, _ in DOC_SLOTS] + [k for k, _ in TEST_SLOTS]


def _next_number():
    """4M 변경 신고서 번호 — 4M-YYYY-NNNN"""
    year = datetime.now().year
    n = DesignChange.query.filter(DesignChange.change_number.like(f'4M-{year}-%')).count()
    # 구 번호체계(DCR-) 데이터와 중복 방지
    return f'4M-{year}-{n + 1:04d}'


def _parse_date(v):
    return datetime.strptime(v, '%Y-%m-%d').date() if v else None


def _save_attachments(dc, slot):
    """변경 전/후 사진 및 관련자료 첨부 저장"""
    files = request.files.getlist(f'files_{slot}')
    if not files:
        return
    from config import Config
    from werkzeug.utils import secure_filename
    from utils import allowed_file
    for f in files:
        if not f or not f.filename:
            continue
        if not allowed_file(f.filename, Config.ALLOWED_EXTENSIONS):
            flash(f'허용되지 않는 파일 형식: {f.filename}', 'warning')
            continue
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        fname = f"4M_{dc.id}_{slot}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(f.filename)}"
        fpath = os.path.join(Config.UPLOAD_FOLDER, fname)
        f.save(fpath)
        db.session.add(ChangeAttachment(
            change_id=dc.id, slot=slot, file_path=fpath, file_name=f.filename,
            is_image=(ext in IMAGE_EXTS), uploaded_by_id=current_user.id))
        # 항목별 첨부가 등록되면 해당 체크박스를 '첨부'로 표시
        if hasattr(dc, slot) and slot not in ('before', 'after', 'etc'):
            setattr(dc, slot, True)


def _apply_form(dc):
    """4M 변경 신고서(P-038-1) 입력값 반영"""
    f = request.form
    dc.product_name = f.get('product_name')
    dc.product_no = f.get('product_no')
    dc.change_title = f.get('change_title')
    dc.change_reason = f.get('change_reason')
    dc.change_content = f.get('change_content')
    dc.before_spec = f.get('before_spec')
    dc.after_spec = f.get('after_spec')
    # 4M 구분
    dc.m4_man = f.get('m4_man') == 'on'
    dc.m4_machine = f.get('m4_machine') == 'on'
    dc.m4_material = f.get('m4_material') == 'on'
    dc.m4_method = f.get('m4_method') == 'on'
    dc.m4_etc = f.get('m4_etc')
    dc.criteria_key = f.get('criteria_key') or None
    dc.change_type = f.get('change_type')
    # 일정
    dc.request_date = _parse_date(f.get('request_date'))
    dc.sample_date = _parse_date(f.get('sample_date'))
    dc.apply_date = _parse_date(f.get('apply_date'))
    dc.stock_qty = f.get('stock_qty')
    dc.schedule_note = f.get('schedule_note')
    dc.department_id = f.get('department_id', type=int)
    # 관련 서류
    for k in ['doc_inspection', 'doc_process', 'doc_drawing', 'doc_control', 'doc_workstd',
              'test_report', 'test_reliability', 'test_ecn', 'test_risk', 'test_none']:
        setattr(dc, k, f.get(k) == 'on')
    # 관련 서류: 파일 첨부 시 해당 항목 자동 체크 (첨부/해당없음)


@design_bp.route('/')
@login_required
def index():
    status = request.args.get('status', '')
    change_type = request.args.get('type', '')
    q = request.args.get('q', '')

    query = DesignChange.query
    if status:
        query = query.filter_by(status=status)
    if change_type:
        query = query.filter_by(change_type=change_type)
    if q:
        query = query.filter(
            DesignChange.product_name.contains(q) |
            DesignChange.change_title.contains(q) |
            DesignChange.change_number.contains(q)
        )

    changes = query.order_by(DesignChange.created_at.desc()).all()
    stats = {
        'draft': DesignChange.query.filter_by(status='draft').count(),
        'review': DesignChange.query.filter_by(status='review').count(),
        'approved': DesignChange.query.filter_by(status='approved').count(),
        'implemented': DesignChange.query.filter_by(status='implemented').count(),
    }
    return render_template('design/index.html',
        changes=changes, stats=stats,
        status=status, change_type=change_type, q=q)


@design_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not current_user.has_edit():
        flash('4M 변경 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('design.index'))
    from routes.m4_data import CRITERIA, REVIEW_DEPTS
    if request.method == 'POST':
        dc = DesignChange(change_number=_next_number(),
                          requester_id=current_user.id, status='draft')
        _apply_form(dc)
        if not dc.request_date:
            dc.request_date = date.today()
        db.session.add(dc)
        db.session.flush()
        for slot in ALL_SLOTS:
            _save_attachments(dc, slot)
        # 관계부서 검토 슬롯 생성 (절차서 5.2.1)
        for d in request.form.getlist('review_depts') or REVIEW_DEPTS:
            db.session.add(ChangeReview(change_id=dc.id, dept_label=d))
        db.session.add(AuditTrail(user_id=current_user.id, action='4M변경등록',
                                  target_type='design_change', target_id=dc.id,
                                  target_name=f'{dc.change_number} {dc.change_title}'))
        db.session.commit()
        flash(f'4M 변경 신고서 [{dc.change_number}]이(가) 등록되었습니다. 관계부서 검토를 요청하세요.', 'success')
        return redirect(url_for('design.detail', dcid=dc.id))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('design/new.html', departments=departments, today=date.today(),
                           item=None, criteria=CRITERIA, review_depts=REVIEW_DEPTS,
                           doc_slots=DOC_SLOTS, test_slots=TEST_SLOTS)


@design_bp.route('/<int:dcid>/edit', methods=['GET', 'POST'])
@login_required
def edit(dcid):
    dc = DesignChange.query.get_or_404(dcid)
    if not current_user.has_edit() or not (current_user.is_admin() or dc.requester_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    if dc.status not in ('draft', 'review'):
        flash('품질검토 이후 단계는 수정할 수 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    from routes.m4_data import CRITERIA, REVIEW_DEPTS
    if request.method == 'POST':
        _apply_form(dc)
        for slot in ALL_SLOTS:
            _save_attachments(dc, slot)
        db.session.add(AuditTrail(user_id=current_user.id, action='4M변경수정',
            target_type='design_change', target_id=dc.id,
            target_name=f'{dc.change_number} {dc.change_title}'))
        db.session.commit()
        flash(f'4M 변경 신고서 [{dc.change_number}]이(가) 수정되었습니다.', 'success')
        return redirect(url_for('design.detail', dcid=dc.id))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('design/new.html', departments=departments, today=date.today(),
                           item=dc, criteria=CRITERIA, review_depts=REVIEW_DEPTS,
                           doc_slots=DOC_SLOTS, test_slots=TEST_SLOTS)


@design_bp.route('/<int:dcid>/submit', methods=['POST'])
@login_required
def submit(dcid):
    """관계부서 검토 요청 (절차서 5.2)"""
    dc = DesignChange.query.get_or_404(dcid)
    if not (current_user.is_admin() or dc.requester_id == current_user.id):
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    dc.status = 'review'
    db.session.add(AuditTrail(user_id=current_user.id, action='4M부서검토요청',
        target_type='design_change', target_id=dc.id, target_name=dc.change_number))
    db.session.commit()
    flash('관계부서 검토가 요청되었습니다.', 'success')
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/<int:dcid>/review/<int:rv_id>', methods=['POST'])
@login_required
def review(dcid, rv_id):
    """관계부서 검토의견 등록·서명 (절차서 5.2.2)"""
    dc = DesignChange.query.get_or_404(dcid)
    rv = ChangeReview.query.get_or_404(rv_id)
    if not current_user.has_review():
        flash('검토 권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    rv.reviewer_name = request.form.get('reviewer_name') or current_user.name
    rv.opinion = request.form.get('opinion')
    rv.signed_by_id = current_user.id
    rv.signed_at = datetime.utcnow()
    db.session.add(AuditTrail(user_id=current_user.id, action='4M부서검토',
        target_type='design_change', target_id=dc.id,
        target_name=f'{dc.change_number} ({rv.dept_label})'))
    db.session.commit()
    flash(f'{rv.dept_label} 검토의견이 등록되었습니다.', 'success')
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/<int:dcid>/qa', methods=['POST'])
@login_required
def qa_review(dcid):
    """품질관리팀 검토·승인 (절차서 5.3)"""
    import json
    dc = DesignChange.query.get_or_404(dcid)
    if not current_user.has_approve():
        flash('품질검토 승인 권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    from routes.m4_data import QA_EVAL_ITEMS
    f = request.form
    dc.qa_receipt_date = _parse_date(f.get('qa_receipt_date')) or dc.qa_receipt_date or date.today()
    if not dc.qa_receiver_id:
        dc.qa_receiver_id = current_user.id
    dc.qa_eval_needed = f.get('qa_eval_needed')
    dc.qa_eval_reason = f.get('qa_eval_reason')
    items = {}
    for i, name in enumerate(QA_EVAL_ITEMS):
        items[name] = {
            'need': f.get('ev_need_%d' % i) or '',
            'date': f.get('ev_date_%d' % i) or '',
            'result': f.get('ev_result_%d' % i) or '',
        }
    dc.qa_eval_items = json.dumps(items, ensure_ascii=False)
    dc.qa_review_date = _parse_date(f.get('qa_review_date')) or date.today()
    dc.qa_reviewer_id = current_user.id
    dc.qa_result = f.get('qa_result')
    dc.qa_reject_reason = f.get('qa_reject_reason')
    dc.customer_target = f.get('customer_target')
    dc.customer_reason = f.get('customer_reason')
    dc.customer_name = f.get('customer_name')

    if dc.qa_result == '불합격':
        dc.status = 'rejected'
        msg = '품질검토 결과 불합격 처리되었습니다.'
    elif dc.customer_target == '대상':
        dc.status = 'customer'
        msg = '품질검토 합격 — 고객사 승인 절차를 진행하세요.'
    else:
        dc.status = 'approved'
        dc.approver_id = current_user.id
        dc.approved_date = date.today()
        msg = '품질검토 합격 — 내부 승인으로 종결되었습니다. (고객통보 불필요)'
    db.session.add(AuditTrail(user_id=current_user.id, action='4M품질검토',
        target_type='design_change', target_id=dc.id,
        target_name=f'{dc.change_number} / {dc.qa_result}'))
    db.session.commit()
    flash(msg, 'success' if dc.qa_result == '합격' else 'warning')
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/<int:dcid>/customer', methods=['POST'])
@login_required
def customer_approve(dcid):
    """고객사 승인 결과 등록 (절차서 5.3.3)"""
    dc = DesignChange.query.get_or_404(dcid)
    if not current_user.has_approve():
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    dc.customer_name = request.form.get('customer_name') or dc.customer_name
    dc.customer_approve_date = _parse_date(request.form.get('customer_approve_date')) or date.today()
    dc.status = 'approved'
    dc.approver_id = current_user.id
    dc.approved_date = date.today()
    db.session.add(AuditTrail(user_id=current_user.id, action='4M고객승인',
        target_type='design_change', target_id=dc.id, target_name=dc.change_number))
    db.session.commit()
    flash('고객사 승인이 등록되어 변경이 승인 완료되었습니다.', 'success')
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/<int:dcid>/implement', methods=['POST'])
@login_required
def implement(dcid):
    """양산 적용(반영) 완료 처리 (절차서 5.4)"""
    dc = DesignChange.query.get_or_404(dcid)
    if not current_user.has_approve():
        flash('권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    dc.status = 'implemented'
    dc.distribution = request.form.get('distribution')
    db.session.add(AuditTrail(user_id=current_user.id, action='4M반영완료',
        target_type='design_change', target_id=dc.id, target_name=dc.change_number))
    db.session.commit()
    flash('변경사항이 반영 완료로 처리되었습니다.', 'success')
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/attachment/<int:att_id>/delete', methods=['POST'])
@login_required
def delete_attachment(att_id):
    att = ChangeAttachment.query.get_or_404(att_id)
    dc = DesignChange.query.get_or_404(att.change_id)
    if not current_user.has_edit() or dc.status in ('approved', 'implemented'):
        flash('첨부 삭제 권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dc.id))
    try:
        if att.file_path and os.path.exists(att.file_path):
            os.remove(att.file_path)
    except OSError:
        pass
    db.session.delete(att)
    db.session.commit()
    flash('첨부가 삭제되었습니다.', 'info')
    return redirect(url_for('design.detail', dcid=dc.id))


@design_bp.route('/<int:dcid>')
@login_required
def detail(dcid):
    dc = DesignChange.query.get_or_404(dcid)
    file_url = None
    if dc.file_path:
        fname = os.path.basename(dc.file_path)
        file_url = url_for('static', filename=f'uploads/{fname}')
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    import json
    from routes.m4_data import (CRITERIA_MAP, DEPT_CHECKPOINTS, QA_EVAL_ITEMS,
                                STATUS_KR, REVIEW_DEPTS)
    reviews = dc.reviews.order_by(ChangeReview.id).all()
    atts = {'before': [], 'after': [], 'etc': []}
    for a in dc.attachments.all():
        atts.setdefault(a.slot or 'etc', []).append(a)
    eval_items = {}
    if dc.qa_eval_items:
        try:
            eval_items = json.loads(dc.qa_eval_items)
        except ValueError:
            eval_items = {}
    criteria = CRITERIA_MAP.get(dc.criteria_key)
    return render_template('design/detail.html', dc=dc, file_url=file_url, users=users,
                           reviews=reviews, atts=atts, eval_items=eval_items,
                           criteria=criteria, checkpoints=DEPT_CHECKPOINTS,
                           qa_items=QA_EVAL_ITEMS, status_kr=STATUS_KR,
                           review_depts=REVIEW_DEPTS, today=date.today(),
                           doc_slots=DOC_SLOTS, test_slots=TEST_SLOTS)


@design_bp.route('/<int:dcid>/approve', methods=['POST'])
@login_required
def approve(dcid):
    if not current_user.has_approve():
        flash('승인 권한이 없습니다.', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    dc = DesignChange.query.get_or_404(dcid)
    action = request.form.get('action')
    # 본인 요청 건은 본인이 승인/기각 불가 (직무분리)
    if dc.requester_id == current_user.id and not current_user.is_admin() and action in ('approve', 'reject'):
        flash('본인이 요청한 설계변경은 승인/기각할 수 없습니다. (직무분리)', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    if action == 'approve':
        dc.status = 'approved'
        dc.approver_id = current_user.id
        dc.approved_date = date.today()
        flash(f'설계변경 [{dc.change_number}]이(가) 승인되었습니다.', 'success')
    elif action == 'reject':
        dc.status = 'rejected'
        flash(f'설계변경 [{dc.change_number}]이(가) 기각되었습니다.', 'warning')
    elif action == 'implement':
        dc.status = 'implemented'
        flash(f'설계변경 [{dc.change_number}]이(가) 반영완료 처리되었습니다.', 'success')
    elif action == 'submit':
        dc.status = 'review'
        flash(f'설계변경 [{dc.change_number}]이(가) 검토 요청되었습니다.', 'info')
    db.session.commit()
    return redirect(url_for('design.detail', dcid=dcid))


@design_bp.route('/<int:dcid>/delete', methods=['POST'])
@login_required
def delete(dcid):
    from deletion import can_delete, remove_file
    dc = DesignChange.query.get_or_404(dcid)
    if not can_delete(dc, 'requester_id', finalized_statuses=('approved', 'implemented', 'rejected')):
        flash('삭제 권한이 없습니다. (작성자는 승인 전까지만, 관리자만 승인건 삭제 가능)', 'danger')
        return redirect(url_for('design.detail', dcid=dcid))
    remove_file(dc.file_path)
    db.session.add(AuditTrail(user_id=current_user.id, action='설계변경삭제',
        target_type='design_change', target_id=dc.id,
        target_name=f'{dc.change_number} {dc.change_title}'))
    db.session.delete(dc)
    db.session.commit()
    flash(f'설계변경 [{dc.change_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('design.index'))


@design_bp.route('/export-excel')
@login_required
def export_excel():
    changes = DesignChange.query.order_by(DesignChange.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '설계변경 관리 대장'

    hf = PatternFill("solid", fgColor="1a3a5c")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:J1')
    ws['A1'] = '주식회사 누리보이스 설계/개발 변경 관리 대장'
    ws['A1'].font = Font(bold=True, size=14, color="1a3a5c")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['변경번호', '제품명', '변경유형', '변경제목', '변경사유(요약)', '요청일', '적용예정일', '요청자', '승인자', '상태']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = center; c.border = thin

    status_map = {'draft': '초안', 'review': '검토중', 'approved': '승인', 'implemented': '반영완료', 'rejected': '기각'}

    for ri, dc in enumerate(changes, 3):
        row_data = [
            dc.change_number, dc.product_name, dc.change_type or '',
            dc.change_title,
            (dc.change_reason or '')[:60],
            str(dc.request_date) if dc.request_date else '',
            str(dc.apply_date) if dc.apply_date else '',
            dc.requester.name if dc.requester else '',
            dc.approver.name if dc.approver else '',
            status_map.get(dc.status, dc.status),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.alignment = center; cell.border = thin

    for i, w in enumerate([14, 20, 14, 30, 40, 12, 12, 12, 12, 10], 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out,
        download_name=f'설계변경관리대장_{datetime.now().strftime("%Y%m%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
