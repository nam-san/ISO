import io
from datetime import datetime, date

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user

from models import (db, RBAItem, RBAAssessment, RBAResult,
                    Department, AuditTrail, User)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

rba_bp = Blueprint('rba', __name__)


# ════════════════════════════════════════════════════════════
# 삼성 RBA 41대 EHS 항목 마스터 시드 데이터
#   (코드, 대분류코드, 대분류명, 항목명, 배점, 권장 증빙 모듈)
#   domain: B*=안전보건 / C*=환경  |  section: ?H/?E=운영 / ?M=경영시스템
# ════════════════════════════════════════════════════════════
RBA_SEED = [
    # ── BH) 안전보건 운영 ──
    ('BH101','BH1','산업안전보건','산업안전보건 인허가',3,'법규 등록부(인허가)'),
    ('BH102','BH1','산업안전보건','안전한 작업장 설계 및 통제',5,'위험성 평가'),
    ('BH103','BH1','산업안전보건','모성보호',2,'실행기록(모성보호)'),
    ('BH201','BH2','비상사태 대비','소방시설 구비',6,'실행기록(소방점검)'),
    ('BH202','BH2','비상사태 대비','비상 대응 프로그램 수립',3,'기준문서(비상대응절차)'),
    ('BH203','BH2','비상사태 대비','피난로,통로 확보',3,'실행기록(점검)'),
    ('BH204','BH2','비상사태 대비','대피훈련,기록관리',2,'교육훈련(대피훈련)'),
    ('BH301','BH3','산업재해 및 질병','산업재해 및 질병 원인분석,시정조치,기록관리',2,'시정조치(CAPA)'),
    ('BH302','BH3','산업재해 및 질병','응급처치 장비와 구급함 이용',2,'실행기록(점검)'),
    ('BH401','BH4','산업위생','유해공정 작업자 관리',6,'위험성 평가'),
    ('BH501','BH5','육체적 과중업무','육체적 과중업무 평가,통제',2,'위험성 평가'),
    ('BH601','BH6','기계안전','안전보호장치,프로그램 실행',6,'위험성 평가'),
    ('BH701','BH7','식품/위생/기숙사','근로자 생활/위생 시설 인프라(휴게실, 복도 등)',2,'실행기록(점검)'),
    ('BH702','BH7','식품/위생/기숙사','작업장 인프라(화장실, 작업장, 식당, 카페 등)',2,'실행기록(점검)'),
    # ── BM) 안전보건 경영시스템 ──
    ('BM101','BM1','위험성 평가(안전보건)','법률/고객 요구사항 준수 프로세스',1,'법규 등록부'),
    ('BM102','BM1','위험성 평가(안전보건)','실사 프로세스',3,'내부심사'),
    ('BM201','BM2','관리 프로세스(안전보건)','전 임직원 책임/권한 부여',2,'기준문서(직무권한)'),
    ('BM202','BM2','관리 프로세스(안전보건)','정책 및 관리 프로세스',1,'기준문서(정책)'),
    ('BM203','BM2','관리 프로세스(안전보건)','교육 프로세스',2,'교육훈련'),
    ('BM301','BM3','커뮤니케이션(안전보건)','근로자, 이해관계자 커뮤니케이션',1,'커뮤니케이션'),
    ('BM302','BM3','커뮤니케이션(안전보건)','근로자 고충/불만 프로세스',1,'근로자 고충처리(/grievance)'),
    ('BM401','BM4','성과 검토 및 지속 개선(안전보건)','성과 검토 및 지속 개선',1,'경영검토'),
    ('BM402','BM4','성과 검토 및 지속 개선(안전보건)','자가 심사 프로세스',1,'내부심사'),
    ('BM403','BM4','성과 검토 및 지속 개선(안전보건)','시정조치 프로세스',1,'시정조치(CAPA)'),
    # ── CE) 환경 운영 ──
    ('CE101','CE1','환경허가 및 보고','환경 인허가',6,'법규 등록부(환경인허가)'),
    ('CE201','CE2','유해물질','유해물질 처리업체 실사',5,'공급업체 실사(미구축)'),
    ('CE301','CE3','고체폐기물','고체 폐기물 관리 및 처리',3,'환경측면 평가'),
    ('CE401','CE4','대기 배출','대기 배출기준 준수',5,'환경측면 평가/측정기록'),
    ('CE402','CE4','대기 배출','법정 소음기준 준수',1,'실행기록(소음측정)'),
    ('CE501','CE5','수자원 관리','수자원 관리 프로세스',5,'환경측면 평가'),
    ('CE601','CE6','에너지 소비, 온실가스 배출','에너지 소비,온실가스 배출 관리',3,'환경측면 평가'),
    # ── CM) 환경 경영시스템 ──
    ('CM101','CM1','위험성 평가(환경)','법률/고객 요구사항 준수 프로세스',1,'법규 등록부'),
    ('CM102','CM1','위험성 평가(환경)','실사 프로세스',2,'내부심사'),
    ('CM201','CM2','관리 프로세스(환경)','전 임직원 책임/권한 부여',1,'기준문서(직무권한)'),
    ('CM202','CM2','관리 프로세스(환경)','정책 및 관리 프로세스',1,'기준문서(정책)'),
    ('CM203','CM2','관리 프로세스(환경)','교육 프로세스',2,'교육훈련'),
    ('CM301','CM3','커뮤니케이션(환경)','근로자, 이해관계자 커뮤니케이션',1,'커뮤니케이션'),
    ('CM302','CM3','커뮤니케이션(환경)','근로자 고충/불만 프로세스',1,'근로자 고충처리(/grievance)'),
    ('CM401','CM4','성과 검토 및 지속 개선(환경)','성과 검토 및 지속 개선',1,'경영검토'),
    ('CM402','CM4','성과 검토 및 지속 개선(환경)','자가 심사 프로세스',1,'내부심사'),
    ('CM403','CM4','성과 검토 및 지속 개선(환경)','시정조치 프로세스',1,'시정조치(CAPA)'),
]


def seed_rba_items():
    """RBA 41대 항목 마스터 데이터 시드 (최초 1회)."""
    if RBAItem.query.count() > 0:
        return
    for i, (code, ccode, cname, name, score, hint) in enumerate(RBA_SEED):
        domain  = '안전보건' if code[0] == 'B' else '환경'
        section = '경영시스템' if code[1] == 'M' else '운영'
        db.session.add(RBAItem(
            code=code, category_code=ccode, category_name=cname,
            domain=domain, section=section, name=name,
            score=score, evidence_hint=hint, sort_order=i,
        ))
    db.session.commit()


# ── 준수율/위반 집계 계산 ─────────────────────────────────────
def _recalc(assessment):
    results = assessment.results.all()
    applicable = 0
    earned = 0
    pc = mc = nc = 0
    for r in results:
        score = r.item.score if r.item else 0
        if r.result == 'na' or r.result == 'pending':
            continue
        applicable += score
        if r.result == 'conformance':
            earned += score
        elif r.result == 'priority':
            pc += 1
        elif r.result == 'major':
            mc += 1
        elif r.result == 'minor':
            nc += 1
    assessment.applicable_score = applicable
    assessment.total_score = earned
    assessment.conformance_rate = round(earned / applicable * 100, 1) if applicable else 0
    assessment.priority_count = pc
    assessment.major_count = mc
    assessment.minor_count = nc


def _next_number():
    year = datetime.now().year
    last = RBAAssessment.query.filter(
        RBAAssessment.assessment_number.like(f'RBA-{year}-%')
    ).order_by(RBAAssessment.id.desc()).first()
    seq = int(last.assessment_number.split('-')[-1]) + 1 if last else 1
    return f'RBA-{year}-{seq:02d}'


# ── 목록 ─────────────────────────────────────────────────────
@rba_bp.route('/')
@login_required
def index():
    assessments = RBAAssessment.query.order_by(RBAAssessment.assess_date.desc()).all()
    items = RBAItem.query.order_by(RBAItem.sort_order).all()

    # 영역별 항목/배점 요약
    domain_summary = {}
    for it in items:
        key = f'{it.domain}-{it.section}'
        d = domain_summary.setdefault(key, {'count': 0, 'score': 0,
                                            'domain': it.domain, 'section': it.section})
        d['count'] += 1
        d['score'] += it.score

    return render_template('rba/index.html',
        assessments=assessments, items=items,
        domain_summary=domain_summary, total_items=len(items))


# ── 신규 자가진단 회차 ───────────────────────────────────────
@rba_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not current_user.has_edit():
        flash('RBA 자가진단 작성 권한이 없습니다.', 'danger')
        return redirect(url_for('rba.index'))

    if request.method == 'POST':
        assess_date = datetime.strptime(request.form['assess_date'], '%Y-%m-%d').date()
        a = RBAAssessment(
            assessment_number = _next_number(),
            title             = request.form['title'],
            customer          = request.form.get('customer') or '삼성전자',
            assess_date       = assess_date,
            assessor_id       = request.form.get('assessor_id', type=int),
            status            = 'draft',
            created_by_id     = current_user.id,
        )
        db.session.add(a)
        db.session.flush()

        # 41개 항목에 대한 빈 결과 행 자동 생성
        items = RBAItem.query.order_by(RBAItem.sort_order).all()
        for it in items:
            db.session.add(RBAResult(assessment_id=a.id, item_id=it.id, result='pending'))

        db.session.add(AuditTrail(user_id=current_user.id, action='RBA자가진단생성',
                                  target_type='rba_assessment', target_id=a.id,
                                  target_name=a.title))
        db.session.commit()
        flash(f'RBA 자가진단 [{a.assessment_number}]가 생성되었습니다. 항목별로 평가하세요.', 'success')
        return redirect(url_for('rba.detail', assessment_id=a.id))

    today = date.today()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('rba/new.html', today=today, users=users,
                           default_title=f'{today.year}년 EHS 운영 자가진단')


# ── 상세 / 항목 평가 ─────────────────────────────────────────
@rba_bp.route('/<int:assessment_id>', methods=['GET', 'POST'])
@login_required
def detail(assessment_id):
    a = RBAAssessment.query.get_or_404(assessment_id)

    if request.method == 'POST':
        if not current_user.has_edit():
            flash('RBA 자가진단 처리 권한이 없습니다.', 'danger')
            return redirect(url_for('rba.detail', assessment_id=a.id))
        action_type = request.form.get('action_type')

        if action_type == 'save_results':
            for r in a.results.all():
                pre = f'r_{r.id}_'
                if (pre + 'result') in request.form:
                    r.result        = request.form.get(pre + 'result')
                    r.evidence_note = request.form.get(pre + 'evidence')
                    r.finding       = request.form.get(pre + 'finding')
                    r.action        = request.form.get(pre + 'action')
                    dept = request.form.get(pre + 'dept', type=int)
                    r.responsible_dept_id = dept if dept else None
                    if r.result in ('conformance', 'minor', 'major', 'priority', 'na'):
                        r.reviewed_date = date.today()
            _recalc(a)
            db.session.commit()
            flash('평가 결과가 저장되었습니다.', 'success')

        elif action_type == 'complete':
            _recalc(a)
            pending = a.results.filter(RBAResult.result == 'pending').count()
            if pending > 0:
                flash(f'미평가 항목이 {pending}건 있습니다. 모두 평가 후 완료 처리하세요.', 'warning')
            else:
                a.status = 'completed'
                db.session.add(AuditTrail(user_id=current_user.id, action='RBA자가진단완료',
                                          target_type='rba_assessment', target_id=a.id,
                                          target_name=a.title))
                db.session.commit()
                flash(f'자가진단이 완료 처리되었습니다. (준수율 {a.conformance_rate}%)', 'success')
            db.session.commit()

        elif action_type == 'reopen':
            a.status = 'draft'
            db.session.commit()
            flash('자가진단이 작성중 상태로 전환되었습니다.', 'info')

        return redirect(url_for('rba.detail', assessment_id=a.id))

    # GET: 결과를 대분류별로 그룹화
    results = a.results.join(RBAItem).order_by(RBAItem.sort_order).all()
    grouped = {}
    for r in results:
        grouped.setdefault(r.item.category_name, []).append(r)

    departments = Department.query.filter_by(is_active=True).all()

    # 통계
    stats = {
        'conformance': sum(1 for r in results if r.result == 'conformance'),
        'minor':       sum(1 for r in results if r.result == 'minor'),
        'major':       sum(1 for r in results if r.result == 'major'),
        'priority':    sum(1 for r in results if r.result == 'priority'),
        'na':          sum(1 for r in results if r.result == 'na'),
        'pending':     sum(1 for r in results if r.result == 'pending'),
    }

    return render_template('rba/detail.html',
        a=a, grouped=grouped, departments=departments, stats=stats,
        total=len(results))


# ── Excel 출력 ────────────────────────────────────────────────
@rba_bp.route('/<int:assessment_id>/export')
@login_required
def export_excel(assessment_id):
    a = RBAAssessment.query.get_or_404(assessment_id)
    results = a.results.join(RBAItem).order_by(RBAItem.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RBA 자가진단'

    hdr_fill = PatternFill('solid', fgColor='1a3a5c')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin     = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

    # 타이틀
    ws.merge_cells('A1:J1')
    ws['A1'] = f'{a.title}  ({a.assessment_number})'
    ws['A1'].font = Font(bold=True, size=14, color='1a3a5c')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # 요약
    ws.merge_cells('A2:J2')
    ws['A2'] = (f'고객사: {a.customer}   |   진단일: {a.assess_date}   |   '
                f'준수율: {a.conformance_rate}%   |   '
                f'Priority {a.priority_count} / Major {a.major_count} / Minor {a.minor_count}')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 20

    headers = ['코드', '대분류', '영역', '항목명', '배점',
               '판정결과', '권장 증빙', '증빙 설명', '위반/관찰', '개선조치']
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin

    result_kr = {'conformance':'충족','minor':'경미','major':'중대',
                 'priority':'최우선','na':'해당없음','pending':'미평가'}
    result_color = {'conformance':'d1fae5','minor':'fef9c3',
                    'major':'fed7aa','priority':'fecaca','na':'e5e7eb','pending':'ffffff'}

    for i, r in enumerate(results, 4):
        it = r.item
        row = [it.code, it.category_name, it.domain + '/' + it.section, it.name, it.score,
               result_kr.get(r.result, r.result), it.evidence_hint,
               r.evidence_note or '', r.finding or '', r.action or '']
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = center if col in (1, 3, 5, 6) else left
            c.border = thin
            if col == 6:
                c.fill = PatternFill('solid', fgColor=result_color.get(r.result, 'ffffff'))

    widths = [10, 22, 14, 36, 6, 10, 20, 30, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(3, i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    filename = f'RBA자가진단_{a.assessment_number}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@rba_bp.route('/<int:assessment_id>/delete', methods=['POST'])
@login_required
def delete(assessment_id):
    from deletion import can_delete
    from models import AuditTrail
    a = RBAAssessment.query.get_or_404(assessment_id)
    if not can_delete(a, 'created_by_id', finalized_statuses=('completed', 'approved')):
        flash('삭제 권한이 없습니다. (작성자는 완료 전까지만, 관리자만 완료건 삭제 가능)', 'danger')
        return redirect(url_for('rba.detail', assessment_id=assessment_id))
    db.session.add(AuditTrail(user_id=current_user.id, action='RBA삭제',
        target_type='rba', target_id=a.id, target_name=a.assessment_number))
    db.session.delete(a)   # 항목결과(results) cascade 삭제
    db.session.commit()
    flash(f'RBA 자가진단 [{a.assessment_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('rba.index'))
