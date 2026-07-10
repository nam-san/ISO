from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Training, TrainingAttendee, TrainingPlan, Department, User, AuditTrail
from datetime import datetime, date
import os, io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

training_bp = Blueprint('training', __name__)


def _next_number():
    year = datetime.now().year
    last = Training.query.filter(Training.training_number.like(f'TRN-{year}-%')).count()
    return f'TRN-{year}-{last + 1:04d}'


def _can_manage():
    """교육훈련 계획·실행 관리 권한 (관리자 + 교육훈련 권한 부여자)."""
    return current_user.has_training_edit()


# ── 연간 교육계획 기본 시드 — 2026 연간교육훈련계획서(NV-P-009) 기준 ──
# 구분: 안전보건 / 환경 / 일반법정 / 품질 / 직무일반 / 기타
# (구분, 주관, 교육과정명, 대상, 인원, 시간, 실시주기, 교육방법/기관, 계획월CSV)
TRAINING_PLAN_SEED = [
    # ── 안전보건 (법정·내부·훈련) ──
    ('안전보건', '관리부', '산업안전보건교육(정기)', '사무직·생산직 전 근로자', None, 6, '분기 1회', '사내 자체', '3,6,9,12'),
    ('안전보건', '관리부', '관리감독자 안전보건교육', '관리감독자(팀장·파트장)', None, 16, '연 1회(분할)', '사외 위탁', '4,5'),
    ('안전보건', '관리부', '신규채용자 안전보건교육', '신규 채용자', None, 8, '수시(채용시)', '사내 자체', ''),
    ('안전보건', '안전보건관리책임자', 'MSDS(물질안전보건자료) 교육', '화학물질(플럭스·세척제) 취급자', None, 1, '연 1회 + 수시', '사내 자체', '5'),
    ('안전보건', '안전보건관리책임자', '고위험작업 특별교육(밀폐공간·LOTO 등)', '고위험작업 종사자', None, None, '연 1회 + 수시', '사내 자체', '6'),
    ('안전보건', '안전보건관리책임자', '비상대응조직 역할별 교육', '비상대응팀(지휘·방호복구·의무지원)', None, None, '연 1회 이상', '사내 자체', '4'),
    ('안전보건', '안전보건관리책임자', '비상대응 훈련(화재·대피)', '전 근로자', None, 4, '연 2회 중 1회', '사내(도상+대피훈련)', '5'),
    ('안전보건', '안전보건관리책임자', '비상대응 훈련(화학물질누출·감전 등)', '전 근로자·해당공정', None, 4, '연 2회 중 1회', '사내 자체', '11'),
    ('안전보건', '안전보건관리책임자', '응급처치(심폐소생술 등) 교육', '응급처치 담당자(부서별 1인↑)', None, 4, '연 1회', '사외 위탁', '9'),
    ('안전보건', '안전보건관리책임자', '근골격계 예방교육', '생산직 근로자', None, 1, '연 1회', '사내 자체', '10'),
    ('안전보건', '관리부', '모성보호 교육', '관리감독자, 인사담당자', None, 1, '연 1회', '사내 자체', '3'),
    ('안전보건', '안전보건관리책임자', '위험기계·기구 방호장치 교육', '설비담당자, 관리감독자', None, 2, '연 1회', '사내 자체', '6'),
    # ── 환경 ──
    ('환경', '관리부', '환경법규 및 환경측면 교육', '환경 관련 업무 담당자', None, 2, '연 1회', '사내 자체', '8'),
    # ── 일반(법정) ──
    ('일반법정', '관리부', '성희롱 예방교육', '전 근로자', None, 1, '연 1회', '온라인', '6'),
    ('일반법정', '관리부', '장애인 인식개선교육', '전 근로자', None, 1, '연 1회', '온라인', '3'),
    ('일반법정', '관리부', '개인정보보호교육', '개인정보 담당자·접근직원', None, 1, '연 1회', '온라인', '9'),
    ('일반법정', '관리부', '퇴직급여제도 교육', '연금가입 근로자', None, 1, '연 1회', '온라인', '11'),
    ('일반법정', '관리부', '직장 내 괴롭힘 예방교육', '전 근로자', None, 1, '연 1회', '온라인', '6'),
    # ── 품질/자격 ──
    ('품질', '품질경영부', '3정5S 교육', '전 근로자(생산직 위주)', None, 1, '연 1회', '사내 자체', '1'),
    ('품질', '품질경영부', '내부심사원 양성교육', '내부심사원 지정(예정)자', None, None, '신규 지정 시', '사외 위탁', '3'),
    ('품질', '품질경영부', 'RBA/고객 요구사항 자가심사 교육', '내부심사원, 관리감독자', None, 2, '연 1회(정기심사 전)', '사내 자체', '9'),
    ('품질', '품질경영부', '검사원·특별공정작업자 적격성 평가', '검사원·시험원·특별공정 작업자', None, None, '2년마다(해당자)', '사내 자체 평가', '2'),
    ('품질', '품질경영부', '신제품 교육', '전 근로자', None, 2, '발생시', '사내 교육', ''),
    ('품질', '품질경영부', '고객불만 및 주요이슈 교육', '전 근로자', None, 2, '월 1회', '사내 교육', '1,2,3,4,5,6,7,8,9,10,11,12'),
    ('품질', '품질경영부', '4M 변경사항 교육', '전 근로자', None, 2, '발생시', '사내 교육', ''),
    # ── 직무·일반 ──
    ('직무일반', '관리부', '직무 교육', '전 근로자', None, 20, '연 1회', '사외 위탁', '10'),
    ('직무일반', '관리부', '보안 교육', '전 근로자', None, 1, '연 1회', '온라인', '11'),
    # ── 도급/기타 ──
    ('기타', '안전보건관리책임자', '도급업체 안전보건 정보제공 및 협의체', '사내 상시 도급업체 근로자', None, 1, '월 1회', '사내(당사 주관)', '1,2,3,4,5,6,7,8,9,10,11,12'),
    ('기타', '관리부', '법적/고객사 요청 교육', '전 근로자', None, None, '발생시', '요청사항에 따른 진행', ''),
]


def seed_training_plans():
    """당해 연도 연간 교육계획 기본 항목 시드 (최신 형식이 아니면 재시드)."""
    year = date.today().year
    defaults = TrainingPlan.query.filter_by(year=year, is_default=True).all()
    # 최신 시드(주관 채워짐 & 개수 일치)면 통과
    if defaults and all(d.host for d in defaults) and len(defaults) == len(TRAINING_PLAN_SEED):
        return
    # 구버전/불일치 → 실행이력 없는 기본항목만 제거 후 재시드
    for d in defaults:
        if d.executions.count() == 0:
            db.session.delete(d)
    db.session.commit()
    for cat, host, title, target, cnt, hours, cycle, inst, months in TRAINING_PLAN_SEED:
        db.session.add(TrainingPlan(
            year=year, category=cat, host=host, title=title,
            target_desc=target, target_count=cnt, planned_hours=hours,
            cycle=cycle, institution=inst, plan_months=months,
            is_default=True, is_active=True))
    db.session.commit()
    print(f'[교육계획] {year}년 기본 교육항목 {len(TRAINING_PLAN_SEED)}건 시드 완료')


CAT_ORDER = {'안전보건': 0, '환경': 1, '일반법정': 2, '품질': 3, '직무일반': 4, '기타': 5}
CAT_LIST = ['안전보건', '환경', '일반법정', '품질', '직무일반', '기타']


def _month_cells(plan, year, today):
    """플랜의 월별 상태 매트릭스 {월: (state, exec_id)} 계산.
    state: 'done'(완료) / 'overdue'(미실시) / 'plan'(계획) / 'adhoc'(발생시) / None(빈칸).
    - 계획 월이 있는 교육: 완료/미실시/계획
    - 계획 월이 없는 '발생시' 교육: 현재월 이후에만 회색 '발생시' 버튼, 지난 달(미실적)은 칸 삭제
    """
    execs = {}
    for t in plan.executions.all():
        if t.training_date and t.training_date.year == year:
            execs.setdefault(t.training_date.month, t.id)
    planned = set(plan.planned_month_list)
    as_needed = not planned      # 계획 월 미지정 = 발생시 교육
    cells = {}
    for m in range(1, 13):
        is_past = (year < today.year) or (year == today.year and m < today.month)
        if m in execs:
            cells[m] = ('done', execs[m])
        elif as_needed:
            # 발생시: 현재월 이후(당해연도) 또는 미래연도만 버튼 유지, 지난 달은 삭제
            cells[m] = (None, None) if is_past else ('adhoc', None)
        elif m in planned:
            cells[m] = ('overdue' if is_past else 'plan', None)
        else:
            cells[m] = (None, None)
    return cells


@training_bp.route('/')
@login_required
def index():
    """연간 교육계획 대장 — 월별 캘린더 매트릭스로 계획·실행현황 표시."""
    this_year = date.today().year
    year = request.args.get('year', type=int) or this_year
    cat = request.args.get('cat', '')
    today = date.today()

    query = TrainingPlan.query.filter_by(year=year, is_active=True)
    if cat:
        query = query.filter_by(category=cat)
    plans = sorted(query.all(), key=lambda p: (CAT_ORDER.get(p.category, 9), p.id))
    for p in plans:
        p.cells = _month_cells(p, year, today)   # 템플릿용 매트릭스 부착

    all_plans = TrainingPlan.query.filter_by(year=year, is_active=True).all()
    # 통계: 계획된 월 슬롯 기준 진행률
    total_slots = done_slots = overdue_slots = 0
    for p in all_plans:
        for m, (state, _) in _month_cells(p, year, today).items():
            if state == 'done': total_slots += 1; done_slots += 1
            elif state == 'overdue': total_slots += 1; overdue_slots += 1
            elif state == 'plan': total_slots += 1
    stats = {
        'total': len(all_plans),
        'done_slots': done_slots,
        'overdue_slots': overdue_slots,
        'plan_slots': total_slots - done_slots - overdue_slots,
        'total_slots': total_slots,
        'legal': sum(1 for p in all_plans if p.category in ('안전보건', '일반법정')),
        'rate': round(done_slots / total_slots * 100) if total_slots else 0,
    }

    years = sorted({y[0] for y in db.session.query(TrainingPlan.year).distinct().all()} | {this_year}, reverse=True)
    departments = Department.query.filter_by(is_active=True).all()
    return render_template('training/index.html',
        plans=plans, stats=stats, year=year, years=years, cat=cat,
        departments=departments, current_year=this_year, today=today)


@training_bp.route('/records')
@login_required
def records():
    """전체 실행이력 목록 (기존 대장 형태)."""
    year = request.args.get('year', type=int) or date.today().year
    q = request.args.get('q', '')
    query = Training.query.filter(Training.training_number.like(f'TRN-{year}-%'))
    if q:
        query = query.filter(Training.title.contains(q) | Training.institution.contains(q))
    trainings = query.order_by(Training.training_date.desc()).all()
    years = sorted({y[0] for y in db.session.query(TrainingPlan.year).distinct().all()} | {date.today().year}, reverse=True)
    return render_template('training/records.html', trainings=trainings, q=q, year=year, years=years)


# ── 연간 교육계획 항목 관리 (교육 담당자) ────────────────────────
def _plan_from_form(plan, form):
    plan.year = form.get('year', type=int) or date.today().year
    plan.category = form.get('category', '일반교육')
    plan.host = form.get('host')
    plan.title = form.get('title')
    plan.legal_basis = form.get('legal_basis')
    plan.iso_standard = form.get('iso_standard') or None
    plan.target_desc = form.get('target_desc')
    plan.cycle = form.get('cycle')
    plan.institution = form.get('institution')
    # 계획 월: 체크박스(months) 우선, 없으면 텍스트(plan_months)
    months = form.getlist('months')
    if months:
        plan.plan_months = ','.join(str(m) for m in sorted(int(x) for x in months if x.isdigit()))
    else:
        plan.plan_months = form.get('plan_months', '')
    ph = form.get('planned_hours')
    plan.planned_hours = float(ph) if ph else None
    plan.target_count = form.get('target_count', type=int)
    plan.department_id = form.get('department_id', type=int)
    plan.note = form.get('note')


@training_bp.route('/plan/new', methods=['GET', 'POST'])
@login_required
def plan_new():
    if not _can_manage():
        flash('교육계획 작성 권한이 없습니다.', 'danger')
        return redirect(url_for('training.index'))
    if request.method == 'POST':
        plan = TrainingPlan(is_default=False, created_by_id=current_user.id)
        _plan_from_form(plan, request.form)
        db.session.add(plan)
        db.session.add(AuditTrail(user_id=current_user.id, action='교육계획등록',
            target_type='training_plan', target_id=0, target_name=plan.title))
        db.session.commit()
        flash(f'교육계획 [{plan.title}]이(가) 추가되었습니다.', 'success')
        return redirect(url_for('training.index', year=plan.year))
    departments = Department.query.filter_by(is_active=True).all()
    return render_template('training/plan_form.html', item=None,
        departments=departments, default_year=date.today().year)


@training_bp.route('/plan/<int:pid>/edit', methods=['GET', 'POST'])
@login_required
def plan_edit(pid):
    plan = TrainingPlan.query.get_or_404(pid)
    if not _can_manage():
        flash('교육계획 수정 권한이 없습니다.', 'danger')
        return redirect(url_for('training.index', year=plan.year))
    if request.method == 'POST':
        _plan_from_form(plan, request.form)
        db.session.commit()
        flash('교육계획이 수정되었습니다.', 'success')
        return redirect(url_for('training.plan_detail', pid=plan.id))
    departments = Department.query.filter_by(is_active=True).all()
    return render_template('training/plan_form.html', item=plan,
        departments=departments, default_year=plan.year)


@training_bp.route('/plan/<int:pid>/delete', methods=['POST'])
@login_required
def plan_delete(pid):
    plan = TrainingPlan.query.get_or_404(pid)
    if not (current_user.is_admin() or (_can_manage() and plan.created_by_id == current_user.id)):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('training.plan_detail', pid=pid))
    # 연결된 실행이력은 보존하고 계획 연결만 해제
    for t in plan.executions.all():
        t.plan_id = None
    db.session.add(AuditTrail(user_id=current_user.id, action='교육계획삭제',
        target_type='training_plan', target_id=plan.id, target_name=plan.title))
    year = plan.year
    db.session.delete(plan)
    db.session.commit()
    flash('교육계획이 삭제되었습니다. (실행이력은 보존)', 'info')
    return redirect(url_for('training.index', year=year))


@training_bp.route('/plan/<int:pid>')
@login_required
def plan_detail(pid):
    plan = TrainingPlan.query.get_or_404(pid)
    executions = plan.executions.order_by(Training.training_date.desc()).all()
    return render_template('training/plan_detail.html', plan=plan, executions=executions)


def _apply_fields(training, form):
    """폼 → Training 기본 필드 반영 (등록·수정 공용)."""
    td = form.get('training_date'); ed = form.get('end_date')
    training.title = form.get('title')
    training.training_type = form.get('training_type')
    training.iso_standard = form.get('iso_standard')
    training.training_date = datetime.strptime(td, '%Y-%m-%d').date() if td else date.today()
    training.end_date = datetime.strptime(ed, '%Y-%m-%d').date() if ed else None
    training.hours = float(form.get('hours', 0) or 0)
    training.institution = form.get('institution')
    training.location = form.get('location')
    training.content = form.get('content')
    training.department_id = form.get('department_id', type=int)
    pid = form.get('plan_id', type=int)
    if pid:
        training.plan_id = pid


def _save_upload(training):
    """첨부파일 저장. 허용되지 않는 형식이면 False."""
    file = request.files.get('file')
    if file and file.filename:
        from config import Config
        from werkzeug.utils import secure_filename
        from utils import allowed_file
        if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
            return False
        fname = f"TRN_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
        fpath = os.path.join(Config.UPLOAD_FOLDER, fname)
        file.save(fpath)
        training.file_path = fpath
        training.file_name = file.filename
    return True


def _save_attendees(training, form):
    """참석자 저장 — 전사(인원수)/개별(이름) 모드. 기존 개별 명단은 교체."""
    mode = form.get('attendee_mode', 'individual')
    training.attendee_mode = mode
    TrainingAttendee.query.filter_by(training_id=training.id).delete()
    if mode == 'company':
        training.headcount = form.get('headcount', type=int) or 0
    else:
        training.headcount = None
        dept_name = form.get('attendee_dept', '')
        for name in [n.strip() for n in form.get('attendee_names', '').split(',') if n.strip()]:
            user = User.query.filter_by(name=name).first()
            db.session.add(TrainingAttendee(
                training_id=training.id, user_id=user.id if user else None,
                name=name, department_name=dept_name, completed=True))


@training_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not _can_manage():
        flash('교육훈련 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('training.index'))
    if request.method == 'POST':
        training = Training(training_number=_next_number(), created_by_id=current_user.id)
        _apply_fields(training, request.form)
        if not _save_upload(training):
            flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
            return redirect(request.url)
        db.session.add(training)
        db.session.flush()
        _save_attendees(training, request.form)
        db.session.add(AuditTrail(user_id=current_user.id, action='교육등록',
            target_type='training', target_id=training.id, target_name=training.title))
        db.session.commit()
        flash(f'교육훈련 [{training.training_number}] {training.title}이(가) 등록되었습니다.', 'success')
        if training.plan_id:
            return redirect(url_for('training.plan_detail', pid=training.plan_id))
        return redirect(url_for('training.records'))

    departments = Department.query.filter_by(is_active=True).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    # 계획/월에서 실행 등록으로 진입한 경우 프리필
    plan = TrainingPlan.query.get(request.args.get('plan_id', type=int)) if request.args.get('plan_id') else None
    month = request.args.get('month', type=int)
    yr = request.args.get('year', type=int) or date.today().year
    try:
        default_date = date(yr, month, 1) if month else date.today()
    except ValueError:
        default_date = date.today()
    return render_template('training/new.html', departments=departments, users=users,
        today=date.today(), default_date=default_date, item=None, plan=plan)


@training_bp.route('/<int:tid>/edit', methods=['GET', 'POST'])
@login_required
def edit(tid):
    training = Training.query.get_or_404(tid)
    if not _can_manage() or not (current_user.is_admin() or training.created_by_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('training.detail', tid=tid))
    if request.method == 'POST':
        _apply_fields(training, request.form)
        if not _save_upload(training):
            flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
            return redirect(request.url)
        _save_attendees(training, request.form)
        db.session.add(AuditTrail(user_id=current_user.id, action='교육수정',
            target_type='training', target_id=training.id, target_name=training.title))
        db.session.commit()
        flash(f'교육훈련 [{training.training_number}] 정보가 수정되었습니다.', 'success')
        return redirect(url_for('training.detail', tid=tid))

    departments = Department.query.filter_by(is_active=True).all()
    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('training/new.html', departments=departments, users=users,
        today=date.today(), default_date=date.today(), item=training, plan=training.plan)


@training_bp.route('/<int:tid>')
@login_required
def detail(tid):
    t = Training.query.get_or_404(tid)
    attendees = TrainingAttendee.query.filter_by(training_id=tid).all()
    file_url = None
    if t.file_path:
        import os as _os
        fname = _os.path.basename(t.file_path)
        file_url = url_for('static', filename=f'uploads/{fname}')
    return render_template('training/detail.html', t=t, attendees=attendees, file_url=file_url)


@training_bp.route('/<int:tid>/delete', methods=['POST'])
@login_required
def delete(tid):
    from deletion import can_delete, remove_file
    t = Training.query.get_or_404(tid)
    if not can_delete(t, 'created_by_id'):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('training.detail', tid=tid))
    remove_file(t.file_path)
    db.session.add(AuditTrail(user_id=current_user.id, action='교육삭제',
        target_type='training', target_id=t.id, target_name=t.title))
    db.session.delete(t)
    db.session.commit()
    flash(f'교육훈련 [{t.training_number}] 기록이 삭제되었습니다.', 'info')
    return redirect(url_for('training.index'))


@training_bp.route('/export-excel')
@login_required
def export_excel():
    year = request.args.get('year', type=int)
    q = Training.query
    if year:
        q = q.filter(Training.training_number.like(f'TRN-{year}-%'))
    trainings = q.order_by(Training.training_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '교육훈련 이수 대장'

    hf = PatternFill("solid", fgColor="1a3a5c")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:J1')
    ws['A1'] = '주식회사 누리보이스 교육훈련 이수 대장'
    ws['A1'].font = Font(bold=True, size=14, color="1a3a5c")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['교육번호', '교육명', '교육유형', 'ISO규격', '교육일', '교육기관', '교육시간(h)', '부서', '참석인원', '첨부']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = center; c.border = thin

    for ri, t in enumerate(trainings, 3):
        att_count = f'{t.attendee_total}명(전사)' if t.attendee_mode == 'company' else t.attendee_total
        row_data = [
            t.training_number, t.title, t.training_type or '', t.iso_standard or '',
            str(t.training_date) if t.training_date else '',
            t.institution or '', t.hours or 0,
            t.department.name if t.department else '',
            att_count, '있음' if t.file_path else ''
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment = center; c.border = thin

    for i, w in enumerate([14, 30, 12, 12, 12, 20, 10, 15, 8, 6], 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out,
        download_name=f'교육훈련이수대장_{datetime.now().strftime("%Y%m%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
