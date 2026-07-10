from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, CustomerComplaint, ComplaintAttachment, User, AuditTrail
from datetime import datetime, date
import io, os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

complaint_bp = Blueprint('complaint', __name__)


def _next_number():
    year = datetime.now().year
    last = CustomerComplaint.query.filter(CustomerComplaint.complaint_number.like(f'CLM-{year}-%')).count()
    return f'CLM-{year}-{last + 1:04d}'


@complaint_bp.route('/')
@login_required
def index():
    status = request.args.get('status', '')
    priority = request.args.get('priority', '')
    q = request.args.get('q', '')

    query = CustomerComplaint.query
    if status:
        query = query.filter_by(status=status)
    if priority:
        query = query.filter_by(priority=priority)
    if q:
        query = query.filter(
            CustomerComplaint.customer_name.contains(q) |
            CustomerComplaint.complaint_content.contains(q) |
            CustomerComplaint.complaint_number.contains(q)
        )

    complaints = query.order_by(CustomerComplaint.receipt_date.desc()).all()

    # 현황 통계
    today = date.today()
    stats = {
        'open': CustomerComplaint.query.filter_by(status='open').count(),
        'investigating': CustomerComplaint.query.filter_by(status='investigating').count(),
        'overdue': CustomerComplaint.query.filter(
            CustomerComplaint.due_date < today,
            CustomerComplaint.status.in_(['open', 'investigating'])
        ).count(),
    }

    return render_template('complaint/index.html',
        complaints=complaints, stats=stats, today=today,
        status=status, priority=priority, q=q)


@complaint_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not current_user.has_edit():
        flash('클레임 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('complaint.index'))
    if request.method == 'POST':
        receipt_date_str = request.form.get('receipt_date')
        due_date_str = request.form.get('due_date')

        c = CustomerComplaint(
            complaint_number=_next_number(),
            receipt_date=datetime.strptime(receipt_date_str, '%Y-%m-%d').date() if receipt_date_str else date.today(),
            customer_name=request.form.get('customer_name'),
            contact_person=request.form.get('contact_person'),
            contact_info=request.form.get('contact_info'),
            product_model=request.form.get('product_model'),
            complaint_type=request.form.get('complaint_type'),
            complaint_content=request.form.get('complaint_content'),
            iso_clause=request.form.get('iso_clause'),
            priority=request.form.get('priority', 'normal'),
            due_date=datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None,
            responsible_id=request.form.get('responsible_id', type=int),
            status='open',
            created_by_id=current_user.id,
        )

        file = request.files.get('file')
        if file and file.filename:
            from config import Config
            from werkzeug.utils import secure_filename
            from utils import allowed_file
            import os
            if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
                return redirect(request.url)
            fname = f"CLM_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            fpath = os.path.join(Config.UPLOAD_FOLDER, fname)
            file.save(fpath)
            c.file_path = fpath
            c.file_name = file.filename

        db.session.add(c)
        db.session.flush()
        log = AuditTrail(user_id=current_user.id, action='클레임등록',
                         target_type='complaint', target_id=c.id,
                         target_name=f'{c.complaint_number} {c.customer_name}')
        db.session.add(log)
        db.session.commit()
        flash(f'고객 클레임 [{c.complaint_number}]이(가) 등록되었습니다.', 'success')
        return redirect(url_for('complaint.detail', cid=c.id))

    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('complaint/new.html', users=users, today=date.today(), item=None)


@complaint_bp.route('/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def edit(cid):
    c = CustomerComplaint.query.get_or_404(cid)
    if not current_user.has_edit() or not (current_user.is_admin() or c.created_by_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    if c.status in ('resolved', 'closed'):
        flash('처리완료·종결된 클레임의 기본정보는 수정할 수 없습니다.', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    if request.method == 'POST':
        receipt_date_str = request.form.get('receipt_date')
        due_date_str = request.form.get('due_date')
        c.receipt_date = datetime.strptime(receipt_date_str, '%Y-%m-%d').date() if receipt_date_str else c.receipt_date
        c.customer_name = request.form.get('customer_name')
        c.contact_person = request.form.get('contact_person')
        c.contact_info = request.form.get('contact_info')
        c.product_model = request.form.get('product_model')
        c.complaint_type = request.form.get('complaint_type')
        c.complaint_content = request.form.get('complaint_content')
        c.iso_clause = request.form.get('iso_clause')
        c.priority = request.form.get('priority', 'normal')
        c.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None
        c.responsible_id = request.form.get('responsible_id', type=int)

        file = request.files.get('file')
        if file and file.filename:
            from config import Config
            from werkzeug.utils import secure_filename
            from utils import allowed_file
            import os
            if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
                return redirect(request.url)
            fname = f"CLM_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            fpath = os.path.join(Config.UPLOAD_FOLDER, fname)
            file.save(fpath)
            c.file_path = fpath
            c.file_name = file.filename

        db.session.add(AuditTrail(user_id=current_user.id, action='클레임수정',
            target_type='complaint', target_id=c.id,
            target_name=f'{c.complaint_number} {c.customer_name}'))
        db.session.commit()
        flash(f'고객 클레임 [{c.complaint_number}]이(가) 수정되었습니다.', 'success')
        return redirect(url_for('complaint.detail', cid=c.id))

    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('complaint/new.html', users=users, today=date.today(), item=c)


@complaint_bp.route('/<int:cid>')
@login_required
def detail(cid):
    c = CustomerComplaint.query.get_or_404(cid)
    file_url = None
    if c.file_path:
        import os
        fname = os.path.basename(c.file_path)
        file_url = url_for('static', filename=f'uploads/{fname}')
    return render_template('complaint/detail.html', c=c, file_url=file_url, today=date.today())


@complaint_bp.route('/<int:cid>/update', methods=['POST'])
@login_required
def update(cid):
    if not current_user.has_edit():
        flash('클레임 처리 권한이 없습니다.', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    c = CustomerComplaint.query.get_or_404(cid)
    if c.status == 'closed':
        flash('종결된 클레임은 더 이상 수정할 수 없습니다.', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    new_status = request.form.get('status')
    c.root_cause = request.form.get('root_cause', c.root_cause)
    c.action_taken = request.form.get('action_taken', c.action_taken)
    c.preventive_measure = request.form.get('preventive_measure', c.preventive_measure)
    if new_status:
        c.status = new_status
        # 처리완료·종결로 전환 시 처리완료일 자동 기입 (최초 완료 시점 보존)
        if new_status in ('resolved', 'closed') and not c.resolved_date:
            c.resolved_date = date.today()
    _save_attachments(c)
    db.session.commit()
    flash('클레임 처리 현황이 업데이트되었습니다.', 'success')
    return redirect(url_for('complaint.detail', cid=cid))


IMAGE_EXTS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}


def _save_attachments(c):
    """처리 증빙 파일들을 ComplaintAttachment로 저장 (이미지는 is_image=True)."""
    files = request.files.getlist('attachments')
    if not files:
        return
    from config import Config
    from werkzeug.utils import secure_filename
    from utils import allowed_file
    for file in files:
        if not file or not file.filename:
            continue
        if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
            flash(f'허용되지 않는 파일 형식: {file.filename}', 'warning')
            continue
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        fname = f"CLM_{c.id}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(file.filename)}"
        fpath = os.path.join(Config.UPLOAD_FOLDER, fname)
        file.save(fpath)
        db.session.add(ComplaintAttachment(
            complaint_id=c.id, file_path=fpath, file_name=file.filename,
            is_image=(ext in IMAGE_EXTS), uploaded_by_id=current_user.id))


@complaint_bp.route('/attachment/<int:att_id>/delete', methods=['POST'])
@login_required
def delete_attachment(att_id):
    att = ComplaintAttachment.query.get_or_404(att_id)
    cid = att.complaint_id
    c = CustomerComplaint.query.get_or_404(cid)
    if not current_user.has_edit():
        flash('삭제 권한이 없습니다.', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    if c.status == 'closed':
        flash('종결된 클레임의 첨부는 삭제할 수 없습니다.', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    try:
        if att.file_path and os.path.exists(att.file_path):
            os.remove(att.file_path)
    except OSError:
        pass
    db.session.delete(att)
    db.session.commit()
    flash('처리 증빙 첨부가 삭제되었습니다.', 'info')
    return redirect(url_for('complaint.detail', cid=cid))


@complaint_bp.route('/<int:cid>/delete', methods=['POST'])
@login_required
def delete(cid):
    from deletion import can_delete, remove_file
    c = CustomerComplaint.query.get_or_404(cid)
    if not can_delete(c, 'created_by_id', finalized_statuses=('resolved', 'closed')):
        flash('삭제 권한이 없습니다. (작성자는 처리완료 전까지만, 관리자만 완료건 삭제 가능)', 'danger')
        return redirect(url_for('complaint.detail', cid=cid))
    remove_file(c.file_path)
    db.session.add(AuditTrail(user_id=current_user.id, action='클레임삭제',
        target_type='complaint', target_id=c.id,
        target_name=f'{c.complaint_number} {c.customer_name}'))
    db.session.delete(c)
    db.session.commit()
    flash(f'고객 클레임 [{c.complaint_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('complaint.index'))


@complaint_bp.route('/export-excel')
@login_required
def export_excel():
    complaints = CustomerComplaint.query.order_by(CustomerComplaint.receipt_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '고객클레임 처리 대장'

    hf = PatternFill("solid", fgColor="1a3a5c")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:K1')
    ws['A1'] = '주식회사 누리보이스 고객클레임 처리 대장'
    ws['A1'].font = Font(bold=True, size=14, color="1a3a5c")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['접수번호', '접수일', '고객사', '제품/모델', '유형', '우선순위', '처리담당자', '처리기한', '완료일', '상태', '처리내용 요약']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = center; c.border = thin

    status_map = {'open': '접수', 'investigating': '조사중', 'resolved': '처리완료', 'closed': '종결'}
    priority_map = {'high': '긴급', 'normal': '일반', 'low': '낮음'}

    for ri, comp in enumerate(complaints, 3):
        row_data = [
            comp.complaint_number,
            str(comp.receipt_date) if comp.receipt_date else '',
            comp.customer_name, comp.product_model or '',
            comp.complaint_type or '',
            priority_map.get(comp.priority, comp.priority),
            comp.responsible.name if comp.responsible else '',
            str(comp.due_date) if comp.due_date else '',
            str(comp.resolved_date) if comp.resolved_date else '',
            status_map.get(comp.status, comp.status),
            (comp.action_taken or '')[:80],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.alignment = center; cell.border = thin
            if comp.status in ('open', 'investigating') and comp.due_date and comp.due_date < date.today():
                cell.fill = PatternFill("solid", fgColor="FFE0E0")

    for i, w in enumerate([14, 12, 20, 18, 12, 8, 12, 12, 12, 10, 40], 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out,
        download_name=f'고객클레임처리대장_{datetime.now().strftime("%Y%m%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
