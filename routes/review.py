import json
import io
from datetime import datetime, date, timedelta

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user

from models import (db, ManagementReview, ManagementReviewAction,
                    CAPA, Audit, CustomerComplaint, DesignChange,
                    Training, LegalRegister, Record, Document, AuditTrail, User)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

review_bp = Blueprint('review', __name__)


# ── 자동채번 ──────────────────────────────────────────────────
def _next_number():
    year = datetime.now().year
    last = ManagementReview.query.filter(
        ManagementReview.review_number.like(f'MR-{year}-%')
    ).order_by(ManagementReview.id.desc()).first()
    seq = int(last.review_number.split('-')[-1]) + 1 if last else 1
    return f'MR-{year}-{seq:02d}'


# ── 현황 데이터 자동 집계 ─────────────────────────────────────
def _build_snapshot(period_start, period_end):
    """검토 기간 내 각 모듈의 핵심 수치를 집계해 dict로 반환."""
    today = date.today()

    # CAPA
    capa_open       = CAPA.query.filter(CAPA.status.in_(['open', 'in_progress'])).count()
    capa_overdue    = CAPA.query.filter(CAPA.status.in_(['open', 'in_progress']),
                                        CAPA.due_date < today).count()
    capa_closed_period = CAPA.query.filter(
        CAPA.completed_date.between(period_start, period_end)).count()

    # 심사
    audit_completed = Audit.query.filter(
        Audit.actual_date.between(period_start, period_end),
        Audit.status == 'completed').count()
    audit_planned   = Audit.query.filter(
        Audit.planned_date.between(period_start, period_end)).count()

    # 고객 클레임
    complaint_total  = CustomerComplaint.query.filter(
        CustomerComplaint.receipt_date.between(period_start, period_end)).count()
    complaint_open   = CustomerComplaint.query.filter(
        CustomerComplaint.status.in_(['open', 'investigating'])).count()
    complaint_high   = CustomerComplaint.query.filter(
        CustomerComplaint.receipt_date.between(period_start, period_end),
        CustomerComplaint.priority == 'high').count()

    # 설계변경
    dcr_total    = DesignChange.query.filter(
        DesignChange.request_date.between(period_start, period_end)).count()
    dcr_approved = DesignChange.query.filter(
        DesignChange.approved_date.between(period_start, period_end)).count()

    # 교육훈련
    training_total   = Training.query.filter(
        Training.training_date.between(period_start, period_end)).count()

    # 법규
    legal_total         = LegalRegister.query.count()
    legal_non_compliant = LegalRegister.query.filter_by(
        compliance_status='non_compliant').count()
    legal_partial       = LegalRegister.query.filter_by(
        compliance_status='partial').count()

    # 기록 보존
    disposal_overdue = Record.query.filter(
        Record.disposal_date < today, Record.status == 'active').count()

    # 문서
    doc_active  = Document.query.filter_by(status='active').count()
    doc_review  = Document.query.filter_by(status='review').count()

    return {
        'period_start': str(period_start),
        'period_end':   str(period_end),
        'capa_open': capa_open,
        'capa_overdue': capa_overdue,
        'capa_closed_period': capa_closed_period,
        'audit_completed': audit_completed,
        'audit_planned': audit_planned,
        'complaint_total': complaint_total,
        'complaint_open': complaint_open,
        'complaint_high': complaint_high,
        'dcr_total': dcr_total,
        'dcr_approved': dcr_approved,
        'training_total': training_total,
        'legal_total': legal_total,
        'legal_non_compliant': legal_non_compliant,
        'legal_partial': legal_partial,
        'disposal_overdue': disposal_overdue,
        'doc_active': doc_active,
        'doc_review': doc_review,
    }


# ── 목록 ─────────────────────────────────────────────────────
@review_bp.route('/')
@login_required
def index():
    reviews = ManagementReview.query.order_by(ManagementReview.review_date.desc()).all()
    # 실행계획 기한 초과 자동 갱신
    today = date.today()
    updated = False
    for r in reviews:
        for a in r.actions.filter(ManagementReviewAction.status == 'pending',
                                  ManagementReviewAction.due_date < today).all():
            a.status = 'overdue'
            updated = True
    if updated:
        db.session.commit()
    return render_template('review/index.html', reviews=reviews, today=today)


# ── 신규 경영검토 ────────────────────────────────────────────
@review_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not current_user.is_manager():
        flash('경영검토 작성 권한이 없습니다.', 'danger')
        return redirect(url_for('review.index'))

    if request.method == 'POST':
        period_start = datetime.strptime(request.form['period_start'], '%Y-%m-%d').date()
        period_end   = datetime.strptime(request.form['period_end'],   '%Y-%m-%d').date()
        review_date  = datetime.strptime(request.form['review_date'],  '%Y-%m-%d').date()
        next_rd_str  = request.form.get('next_review_date')
        next_rd      = datetime.strptime(next_rd_str, '%Y-%m-%d').date() if next_rd_str else None

        snap = _build_snapshot(period_start, period_end)

        mr = ManagementReview(
            review_number   = _next_number(),
            title           = request.form['title'],
            period_start    = period_start,
            period_end      = period_end,
            review_date     = review_date,
            next_review_date= next_rd,
            venue           = request.form.get('venue'),
            chairman_id     = request.form.get('chairman_id', type=int),
            status          = 'draft',
            snapshot_json   = json.dumps(snap, ensure_ascii=False),
            created_by_id   = current_user.id,
        )
        db.session.add(mr)
        db.session.flush()

        log = AuditTrail(user_id=current_user.id, action='경영검토등록',
                         target_type='management_review', target_id=mr.id,
                         target_name=mr.title)
        db.session.add(log)
        db.session.commit()
        flash(f'경영검토 [{mr.review_number}]가 생성되었습니다. 각 안건을 작성하세요.', 'success')
        return redirect(url_for('review.detail', review_id=mr.id))

    # GET — 기본값 계산 (당해연도 상/하반기 자동 선택)
    today = date.today()
    if today.month <= 6:
        default_start = date(today.year, 1, 1)
        default_end   = date(today.year, 6, 30)
        default_title = f'{today.year}년 상반기 경영검토'
    else:
        default_start = date(today.year, 7, 1)
        default_end   = date(today.year, 12, 31)
        default_title = f'{today.year}년 하반기 경영검토'

    users = User.query.filter_by(is_active=True).order_by(User.name).all()
    return render_template('review/new.html',
        today=today, default_start=default_start, default_end=default_end,
        default_title=default_title, users=users)


# ── 상세 / 안건 작성 ──────────────────────────────────────────
@review_bp.route('/<int:review_id>', methods=['GET', 'POST'])
@login_required
def detail(review_id):
    mr = ManagementReview.query.get_or_404(review_id)
    snap = json.loads(mr.snapshot_json) if mr.snapshot_json else {}

    if request.method == 'POST':
        if mr.status == 'approved':
            flash('승인 완료된 경영검토는 수정할 수 없습니다.', 'warning')
            return redirect(url_for('review.detail', review_id=review_id))

        # 경영검토 작성·실행계획 관리는 그룹장/PM/PL 이상 (작성 정책과 동일)
        if not current_user.is_manager():
            flash('경영검토 처리 권한이 없습니다. (그룹장/PM/PL 이상)', 'danger')
            return redirect(url_for('review.detail', review_id=review_id))

        action_type = request.form.get('action_type')

        if action_type == 'save_agenda':
            mr.agenda_prev_actions = request.form.get('agenda_prev_actions')
            mr.agenda_policy       = request.form.get('agenda_policy')
            mr.agenda_audit        = request.form.get('agenda_audit')
            mr.agenda_capa         = request.form.get('agenda_capa')
            mr.agenda_customer     = request.form.get('agenda_customer')
            mr.agenda_process      = request.form.get('agenda_process')
            mr.agenda_supplier     = request.form.get('agenda_supplier')
            mr.agenda_legal        = request.form.get('agenda_legal')
            mr.agenda_risk         = request.form.get('agenda_risk')
            mr.agenda_resources    = request.form.get('agenda_resources')
            mr.agenda_hse          = request.form.get('agenda_hse')
            mr.output_improvements = request.form.get('output_improvements')
            mr.output_changes      = request.form.get('output_changes')
            mr.output_resources    = request.form.get('output_resources')
            mr.minutes             = request.form.get('minutes')
            if mr.status == 'draft':
                mr.status = 'in_review'
            db.session.commit()
            flash('경영검토 내용이 저장되었습니다.', 'success')

        elif action_type == 'add_action':
            due_str = request.form.get('due_date')
            action = ManagementReviewAction(
                review_id      = mr.id,
                category       = request.form.get('category'),
                action_item    = request.form.get('action_item'),
                responsible_id = request.form.get('responsible_id', type=int),
                due_date       = datetime.strptime(due_str, '%Y-%m-%d').date() if due_str else None,
                status         = 'pending',
            )
            db.session.add(action)
            db.session.commit()
            flash('실행계획이 추가되었습니다.', 'success')

        elif action_type == 'update_action':
            action_id = request.form.get('action_id', type=int)
            action = ManagementReviewAction.query.get(action_id)
            if action and action.review_id == mr.id:
                action.status = request.form.get('status')
                action.result = request.form.get('result')
                if action.status == 'completed' and not action.completed_at:
                    action.completed_at = datetime.utcnow()
                db.session.commit()
                flash('실행계획 상태가 업데이트되었습니다.', 'success')

        elif action_type == 'approve':
            if not current_user.is_admin():
                flash('승인 권한이 없습니다.', 'danger')
            else:
                mr.status           = 'approved'
                mr.approved_by_id   = current_user.id
                mr.approved_at      = datetime.utcnow()
                mr.approval_comment = request.form.get('approval_comment')
                log = AuditTrail(user_id=current_user.id, action='경영검토승인',
                                 target_type='management_review', target_id=mr.id,
                                 target_name=mr.title)
                db.session.add(log)
                db.session.commit()
                flash(f'경영검토 [{mr.review_number}]가 승인되었습니다.', 'success')

        elif action_type == 'refresh_snapshot':
            period_start = mr.period_start
            period_end   = mr.period_end
            snap = _build_snapshot(period_start, period_end)
            mr.snapshot_json = json.dumps(snap, ensure_ascii=False)
            db.session.commit()
            flash('현황 데이터가 최신으로 갱신되었습니다.', 'info')

        return redirect(url_for('review.detail', review_id=review_id))

    users   = User.query.filter_by(is_active=True).order_by(User.name).all()
    actions = mr.actions.order_by(ManagementReviewAction.due_date).all()
    today   = date.today()

    action_stats = {
        'total':       len(actions),
        'completed':   sum(1 for a in actions if a.status == 'completed'),
        'in_progress': sum(1 for a in actions if a.status == 'in_progress'),
        'overdue':     sum(1 for a in actions if a.status == 'overdue'),
        'pending':     sum(1 for a in actions if a.status == 'pending'),
    }

    return render_template('review/detail.html',
        mr=mr, snap=snap, actions=actions, users=users,
        action_stats=action_stats, today=today)


# ── Excel 출력 ────────────────────────────────────────────────
@review_bp.route('/<int:review_id>/export')
@login_required
def export_excel(review_id):
    mr   = ManagementReview.query.get_or_404(review_id)
    snap = json.loads(mr.snapshot_json) if mr.snapshot_json else {}

    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──
    hdr_fill  = PatternFill('solid', fgColor='1a3a5c')
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    sub_fill  = PatternFill('solid', fgColor='dbeafe')
    sub_font  = Font(bold=True, color='1e3a5f', size=10)
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap = Alignment(horizontal='left',   vertical='top',    wrap_text=True)
    thin      = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'),  bottom=Side(style='thin'))

    def hdr_cell(ws, row, col, val, span=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + span - 1)
        return c

    def sub_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = sub_fill; c.font = sub_font; c.alignment = left_wrap; c.border = thin
        return c

    def data_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = left_wrap; c.border = thin
        return c

    # ══════════════════════════════════════
    # 시트 1: 경영검토 요약
    # ══════════════════════════════════════
    ws1 = wb.active
    ws1.title = '경영검토 요약'
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 60

    r = 1
    ws1.merge_cells(f'A{r}:B{r}')
    c = ws1.cell(r, 1, '주식회사 누리보이스 경영검토 보고서')
    c.font = Font(bold=True, size=16, color='1a3a5c')
    c.alignment = center; ws1.row_dimensions[r].height = 36

    r += 1
    for label, val in [
        ('문서번호',   mr.review_number),
        ('제목',       mr.title),
        ('검토 대상 기간', f'{mr.period_start} ~ {mr.period_end}'),
        ('경영검토 일시', str(mr.review_date)),
        ('개최 장소',  mr.venue or ''),
        ('의장',       mr.chairman.name if mr.chairman else ''),
        ('상태',       {'draft':'작성중','in_review':'검토중','approved':'승인완료'}.get(mr.status, mr.status)),
        ('승인자',     mr.approved_by.name if mr.approved_by else ''),
        ('승인일시',   mr.approved_at.strftime('%Y-%m-%d %H:%M') if mr.approved_at else ''),
    ]:
        sub_cell(ws1, r, 1, label)
        data_cell(ws1, r, 2, val)
        ws1.row_dimensions[r].height = 18
        r += 1

    r += 1
    agendas = [
        ('전기 후속조치 이행 현황', mr.agenda_prev_actions),
        ('경영방침 및 목표 달성 현황', mr.agenda_policy),
        ('내부심사 결과', mr.agenda_audit),
        ('시정조치(CAPA) 현황', mr.agenda_capa),
        ('고객 피드백 및 클레임', mr.agenda_customer),
        ('프로세스 성과', mr.agenda_process),
        ('외부공급자 성과', mr.agenda_supplier),
        ('법규 준수 현황', mr.agenda_legal),
        ('리스크 및 기회', mr.agenda_risk),
        ('자원 충족성', mr.agenda_resources),
        ('환경·안전(HSE) 성과', mr.agenda_hse),
        ('개선 기회 결정사항 (Output)', mr.output_improvements),
        ('경영시스템 변경사항 (Output)', mr.output_changes),
        ('자원 필요사항 (Output)', mr.output_resources),
        ('종합 회의록', mr.minutes),
    ]
    hdr_cell(ws1, r, 1, '구분'); hdr_cell(ws1, r, 2, '검토 내용'); r += 1
    for label, content in agendas:
        sub_cell(ws1, r, 1, label)
        data_cell(ws1, r, 2, content or '')
        ws1.row_dimensions[r].height = 60
        r += 1

    # ══════════════════════════════════════
    # 시트 2: 현황 스냅샷
    # ══════════════════════════════════════
    ws2 = wb.create_sheet('현황 스냅샷')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 28
    ws2.column_dimensions['D'].width = 15

    ws2.merge_cells('A1:D1')
    c = ws2.cell(1, 1, f'경영검토 현황 스냅샷 ({snap.get("period_start", "")} ~ {snap.get("period_end", "")})')
    c.font = Font(bold=True, size=13, color='1a3a5c'); c.alignment = center
    ws2.row_dimensions[1].height = 28

    snap_rows = [
        ('CAPA 미결 건수',           snap.get('capa_open', 0),           '기한 초과 CAPA',          snap.get('capa_overdue', 0)),
        ('기간 내 CAPA 종결',        snap.get('capa_closed_period', 0),   '완료 심사 건수',          snap.get('audit_completed', 0)),
        ('계획 심사 건수',           snap.get('audit_planned', 0),        '기간 내 클레임 접수',     snap.get('complaint_total', 0)),
        ('미처리 클레임',            snap.get('complaint_open', 0),       '긴급 클레임',             snap.get('complaint_high', 0)),
        ('기간 내 설계변경 요청',    snap.get('dcr_total', 0),            '설계변경 승인',           snap.get('dcr_approved', 0)),
        ('기간 내 교육 실시',        snap.get('training_total', 0),       '법규 등록 건수',          snap.get('legal_total', 0)),
        ('법규 미준수 건수',         snap.get('legal_non_compliant', 0),  '법규 부분준수 건수',      snap.get('legal_partial', 0)),
        ('보존연한 초과 기록',       snap.get('disposal_overdue', 0),     '활성 기준문서 수',        snap.get('doc_active', 0)),
    ]
    r2 = 2
    hdr_cell(ws2, r2, 1, '항목'); hdr_cell(ws2, r2, 2, '수치')
    hdr_cell(ws2, r2, 3, '항목'); hdr_cell(ws2, r2, 4, '수치'); r2 += 1
    for a, b, c_label, d in snap_rows:
        sub_cell(ws2, r2, 1, a); data_cell(ws2, r2, 2, b)
        sub_cell(ws2, r2, 3, c_label); data_cell(ws2, r2, 4, d)
        ws2.row_dimensions[r2].height = 18; r2 += 1

    # ══════════════════════════════════════
    # 시트 3: 실행계획
    # ══════════════════════════════════════
    ws3 = wb.create_sheet('실행계획')
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 40
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 30

    ws3.merge_cells('A1:G1')
    c = ws3.cell(1, 1, f'경영검토 실행계획 — {mr.title}')
    c.font = Font(bold=True, size=13, color='1a3a5c'); c.alignment = center
    ws3.row_dimensions[1].height = 28

    headers = ['No', '구분', '실행항목', '담당자', '기한', '상태', '완료결과/비고']
    for col, h in enumerate(headers, 1):
        hdr_cell(ws3, 2, col, h)

    status_map = {'pending':'대기','in_progress':'진행중','completed':'완료','overdue':'기한초과'}
    for i, a in enumerate(mr.actions.order_by(ManagementReviewAction.due_date).all(), 1):
        row = [i, a.category, a.action_item,
               a.responsible.name if a.responsible else '',
               str(a.due_date) if a.due_date else '',
               status_map.get(a.status, a.status),
               a.result or '']
        for col, val in enumerate(row, 1):
            c = ws3.cell(3 + i - 1, col, val)
            c.alignment = left_wrap; c.border = thin
            if a.status == 'completed':
                c.fill = PatternFill('solid', fgColor='d1fae5')
            elif a.status == 'overdue':
                c.fill = PatternFill('solid', fgColor='fee2e2')
        ws3.row_dimensions[3 + i - 1].height = 30

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    filename = f'경영검토_{mr.review_number}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@review_bp.route('/<int:review_id>/delete', methods=['POST'])
@login_required
def delete(review_id):
    from deletion import can_delete
    from models import AuditTrail
    mr = ManagementReview.query.get_or_404(review_id)
    if not can_delete(mr, 'created_by_id', finalized_statuses=('approved',)):
        flash('삭제 권한이 없습니다. (작성자는 승인 전까지만, 관리자만 승인건 삭제 가능)', 'danger')
        return redirect(url_for('review.detail', review_id=review_id))
    db.session.add(AuditTrail(user_id=current_user.id, action='경영검토삭제',
        target_type='review', target_id=mr.id, target_name=mr.review_number))
    db.session.delete(mr)   # 실행계획(actions) cascade 삭제
    db.session.commit()
    flash(f'경영검토 [{mr.review_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('review.index'))
