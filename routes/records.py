from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, Record, Department, Document, AuditTrail
from datetime import datetime, date, timedelta
import os, io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

records_bp = Blueprint('records', __name__)


def _next_record_number(record_type):
    prefix = 'REC'
    last = Record.query.filter(
        Record.record_number.like(f'{prefix}-%')
    ).order_by(Record.id.desc()).first()
    num = int(last.record_number.split('-')[-1]) + 1 if last else 1
    return f'{prefix}-{datetime.now().year}-{num:04d}'


@records_bp.route('/')
@login_required
def index():
    q = request.args.get('q', '')
    dept_id = request.args.get('dept', '')
    status = request.args.get('status', '')
    iso = request.args.get('iso', '')

    query = Record.query
    if q:
        query = query.filter(Record.title.contains(q) | Record.record_number.contains(q))
    if dept_id:
        query = query.filter_by(department_id=int(dept_id))
    if status:
        query = query.filter_by(status=status)
    if iso:
        query = query.filter_by(iso_standard=iso)

    # 폐기 예정 30일 이내 알림
    today = date.today()
    disposal_soon = Record.query.filter(
        Record.disposal_date.between(today, today + timedelta(days=30)),
        Record.status == 'active'
    ).count()

    records = query.order_by(Record.created_at.desc()).all()
    departments = Department.query.filter_by(is_active=True).all()

    # 현재 부서 객체 (사이드바 강조용)
    current_dept = Department.query.get(int(dept_id)) if dept_id else None

    return render_template('records/index.html',
        records=records,
        departments=departments,
        disposal_soon=disposal_soon,
        current_dept=current_dept,
        today=today,
        q=q, dept_id=dept_id, status=status, iso=iso
    )


@records_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new():
    if not current_user.has_edit():
        flash('기록 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('records.index'))
    if request.method == 'POST':
        record_date_str = request.form.get('record_date')
        record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date() if record_date_str else date.today()
        retention_years = int(request.form.get('retention_years', 3))
        disposal_date = record_date.replace(year=record_date.year + retention_years)

        rec = Record(
            record_number=_next_record_number(request.form.get('record_type')),
            title=request.form.get('title'),
            record_type=request.form.get('record_type'),
            iso_standard=request.form.get('iso_standard'),
            department_id=request.form.get('department_id', type=int),
            created_by_id=current_user.id,
            record_date=record_date,
            retention_years=retention_years,
            disposal_date=disposal_date,
            content=request.form.get('content'),
            status='active',
        )

        file = request.files.get('file')
        if file and file.filename:
            from config import Config
            from werkzeug.utils import secure_filename
            from utils import allowed_file
            if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
                return redirect(request.url)
            filename = f"REC_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            filepath = os.path.join(Config.UPLOAD_FOLDER, filename)
            file.save(filepath)
            rec.file_path = filepath
            rec.file_name = file.filename

        db.session.add(rec)
        db.session.flush()
        log = AuditTrail(user_id=current_user.id, action='기록등록',
                         target_type='record', target_id=rec.id, target_name=rec.title)
        db.session.add(log)
        db.session.commit()
        flash(f'실행기록 [{rec.record_number}]이(가) 등록되었습니다.', 'success')
        return redirect(url_for('records.index'))

    departments = Department.query.filter_by(is_active=True).all()
    documents = Document.query.filter_by(status='active').all()
    return render_template('records/new.html', departments=departments, documents=documents,
                           today=date.today(), item=None)


@records_bp.route('/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def edit(rid):
    rec = Record.query.get_or_404(rid)
    if not current_user.has_edit() or not (current_user.is_admin() or rec.created_by_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('records.index'))
    if rec.status == 'disposed':
        flash('이미 폐기된 기록은 수정할 수 없습니다.', 'danger')
        return redirect(url_for('records.index'))
    if request.method == 'POST':
        record_date_str = request.form.get('record_date')
        rec.record_date = datetime.strptime(record_date_str, '%Y-%m-%d').date() if record_date_str else rec.record_date
        rec.retention_years = int(request.form.get('retention_years', 3))
        rec.disposal_date = rec.record_date.replace(year=rec.record_date.year + rec.retention_years)
        rec.title = request.form.get('title')
        rec.record_type = request.form.get('record_type')
        rec.iso_standard = request.form.get('iso_standard')
        rec.department_id = request.form.get('department_id', type=int)
        rec.content = request.form.get('content')

        file = request.files.get('file')
        if file and file.filename:
            from config import Config
            from werkzeug.utils import secure_filename
            from utils import allowed_file
            if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
                return redirect(request.url)
            filename = f"REC_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            filepath = os.path.join(Config.UPLOAD_FOLDER, filename)
            file.save(filepath)
            rec.file_path = filepath
            rec.file_name = file.filename

        db.session.add(AuditTrail(user_id=current_user.id, action='기록수정',
                                  target_type='record', target_id=rec.id, target_name=rec.title))
        db.session.commit()
        flash(f'실행기록 [{rec.record_number}]이(가) 수정되었습니다.', 'success')
        return redirect(url_for('records.index'))

    departments = Department.query.filter_by(is_active=True).all()
    documents = Document.query.filter_by(status='active').all()
    return render_template('records/new.html', departments=departments, documents=documents,
                           today=date.today(), item=rec)


@records_bp.route('/<int:rid>')
@login_required
def detail(rid):
    """실행기록 상세 — 첨부 파일 미리보기(PDF는 본문 표시)"""
    rec = Record.query.get_or_404(rid)
    file_url = None
    if rec.file_path:
        file_url = url_for('static', filename=f'uploads/{os.path.basename(rec.file_path)}')
    related_doc = Document.query.get(rec.related_doc_id) if rec.related_doc_id else None
    db.session.add(AuditTrail(user_id=current_user.id, action='기록열람',
        target_type='record', target_id=rec.id,
        target_name=f'{rec.record_number} {rec.title}'))
    db.session.commit()
    return render_template('records/detail.html', rec=rec, file_url=file_url,
                           related_doc=related_doc, today=date.today())


@records_bp.route('/<int:rid>/download')
@login_required
def download(rid):
    """기록 첨부 파일 다운로드"""
    rec = Record.query.get_or_404(rid)
    if not rec.file_path or not os.path.exists(rec.file_path):
        flash('첨부 파일을 찾을 수 없습니다.', 'warning')
        return redirect(url_for('records.detail', rid=rid))
    db.session.add(AuditTrail(user_id=current_user.id, action='기록다운로드',
        target_type='record', target_id=rec.id,
        target_name=f'{rec.record_number} {rec.title}'))
    db.session.commit()
    return send_file(rec.file_path, as_attachment=True,
                     download_name=rec.file_name or os.path.basename(rec.file_path))


@records_bp.route('/<int:rid>/delete', methods=['POST'])
@login_required
def delete(rid):
    from deletion import can_delete, remove_file
    rec = Record.query.get_or_404(rid)
    # 작성자는 폐기 전(active)만, 관리자는 전부
    if not can_delete(rec, 'created_by_id', finalized_statuses=('disposed',)):
        flash('삭제 권한이 없습니다. (작성자는 폐기 전까지만, 관리자만 폐기건 삭제 가능)', 'danger')
        return redirect(url_for('records.index'))
    remove_file(rec.file_path)
    db.session.add(AuditTrail(user_id=current_user.id, action='기록삭제',
        target_type='record', target_id=rec.id, target_name=rec.title))
    db.session.delete(rec)
    db.session.commit()
    flash(f'실행기록 [{rec.record_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('records.index'))


@records_bp.route('/export-excel')
@login_required
def export_excel():
    """실행기록 대장 Excel 출력"""
    records = Record.query.order_by(Record.record_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '실행기록 대장'

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1a3a5c")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 타이틀
    ws.merge_cells('A1:J1')
    ws['A1'] = '주식회사 누리보이스 실행기록 대장'
    ws['A1'].font = Font(bold=True, size=14, color="1a3a5c")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # 컬럼 헤더
    headers = ['기록번호', '제목', '기록유형', 'ISO규격', '부서', '기록일',
               '보존연한(년)', '폐기예정일', '상태', '작성자']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # 데이터
    for row_idx, r in enumerate(records, 3):
        row_data = [
            r.record_number, r.title, r.record_type, r.iso_standard,
            r.department.name if r.department else '',
            str(r.record_date) if r.record_date else '',
            r.retention_years,
            str(r.disposal_date) if r.disposal_date else '',
            '활성' if r.status == 'active' else '폐기',
            r.created_by.name if r.created_by else '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.border = thin
            if r.disposal_date and r.disposal_date <= date.today():
                cell.fill = PatternFill("solid", fgColor="FFE0E0")

    # 열 너비 자동조정
    col_widths = [15, 30, 15, 12, 15, 12, 12, 14, 8, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     download_name=f'실행기록대장_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
