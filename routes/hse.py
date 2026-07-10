from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, EnvironmentalAspect, RiskAssessment, Department, LegalRegister, LegalClause, RiskImprovement
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

hse_bp = Blueprint('hse', __name__)

# 공통 엑셀 스타일
_HDR_FILL = PatternFill('solid', fgColor='1a3a5c')
_HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
_TITLE_FONT = Font(bold=True, size=14, color='1a3a5c')
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
_THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
               top=Side(style='thin'), bottom=Side(style='thin'))


def _xlsx_response(wb, filename):
    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return send_file(out, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@hse_bp.route('/')
@login_required
def index():
    env_count = EnvironmentalAspect.query.count()
    env_significant = EnvironmentalAspect.query.filter_by(is_significant=True).count()
    risk_high = RiskAssessment.query.filter_by(risk_grade='상').count()
    risk_count = RiskAssessment.query.count()
    return render_template('hse/index.html',
        env_count=env_count, env_significant=env_significant,
        risk_high=risk_high, risk_count=risk_count)


@hse_bp.route('/environmental')
@login_required
def environmental():
    site = request.args.get('site', '')
    query = EnvironmentalAspect.query
    if site:
        query = query.filter_by(site=site)
    aspects = query.order_by(EnvironmentalAspect.created_at.desc()).all()
    return render_template('hse/environmental.html', aspects=aspects, site=site)


@hse_bp.route('/environmental/new', methods=['GET', 'POST'])
@login_required
def new_environmental():
    if not current_user.has_edit():
        flash('환경측면 평가 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.environmental'))
    if request.method == 'POST':
        frequency = int(request.form.get('frequency', 1))
        severity = int(request.form.get('severity', 1))
        significance = frequency * severity

        year = datetime.now().year
        count = EnvironmentalAspect.query.filter(
            EnvironmentalAspect.aspect_number.like(f'ENV-{year}-%')
        ).count() + 1

        aspect = EnvironmentalAspect(
            aspect_number=f'ENV-{year}-{count:03d}',
            site=request.form.get('site', '제조공장'),
            process=request.form.get('process'),
            aspect=request.form.get('aspect'),
            impact=request.form.get('impact'),
            frequency=frequency,
            severity=severity,
            significance=significance,
            is_significant=significance >= 12,  # 12점 이상 중요 환경측면
            control_measure=request.form.get('control_measure'),
            department_id=request.form.get('department_id', type=int),
            evaluation_date=datetime.strptime(request.form.get('evaluation_date'), '%Y-%m-%d').date()
                            if request.form.get('evaluation_date') else date.today(),
            created_by_id=current_user.id,
        )
        db.session.add(aspect)
        db.session.commit()
        flash('환경측면 영향평가가 등록되었습니다.', 'success')
        return redirect(url_for('hse.environmental'))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('hse/new_environmental.html', departments=departments, today=date.today(), item=None)


@hse_bp.route('/environmental/<int:aid>/edit', methods=['GET', 'POST'])
@login_required
def edit_environmental(aid):
    aspect = EnvironmentalAspect.query.get_or_404(aid)
    if not current_user.has_edit() or not (current_user.is_admin() or aspect.created_by_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('hse.environmental'))
    if request.method == 'POST':
        frequency = int(request.form.get('frequency', 1))
        severity = int(request.form.get('severity', 1))
        aspect.site = request.form.get('site', aspect.site)
        aspect.process = request.form.get('process')
        aspect.aspect = request.form.get('aspect')
        aspect.impact = request.form.get('impact')
        aspect.frequency = frequency
        aspect.severity = severity
        aspect.significance = frequency * severity
        aspect.is_significant = (frequency * severity) >= 12
        aspect.control_measure = request.form.get('control_measure')
        aspect.department_id = request.form.get('department_id', type=int)
        ed = request.form.get('evaluation_date')
        aspect.evaluation_date = datetime.strptime(ed, '%Y-%m-%d').date() if ed else aspect.evaluation_date
        db.session.commit()
        flash(f'환경측면 영향평가 [{aspect.aspect_number}]이(가) 수정되었습니다.', 'success')
        return redirect(url_for('hse.environmental'))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('hse/new_environmental.html', departments=departments, today=date.today(), item=aspect)


@hse_bp.route('/environmental/<int:aid>/delete', methods=['POST'])
@login_required
def delete_environmental(aid):
    from deletion import can_delete
    from models import AuditTrail
    aspect = EnvironmentalAspect.query.get_or_404(aid)
    if not can_delete(aspect, 'created_by_id'):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('hse.environmental'))
    db.session.add(AuditTrail(user_id=current_user.id, action='환경측면삭제',
        target_type='environmental', target_id=aspect.id, target_name=aspect.aspect_number))
    db.session.delete(aspect)
    db.session.commit()
    flash(f'환경측면 [{aspect.aspect_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('hse.environmental'))


@hse_bp.route('/environmental/export')
@login_required
def export_environmental():
    site = request.args.get('site', '')
    query = EnvironmentalAspect.query
    if site:
        query = query.filter_by(site=site)
    aspects = query.order_by(EnvironmentalAspect.site, EnvironmentalAspect.aspect_number).all()
    deptmap = {d.id: d.name for d in Department.query.all()}

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '환경측면 영향평가'
    ws.merge_cells('A1:L1')
    ws['A1'] = '주식회사 누리보이스 환경측면 영향평가 등록부' + (f' ({site})' if site else '')
    ws['A1'].font = _TITLE_FONT; ws['A1'].alignment = _CENTER; ws.row_dimensions[1].height = 30

    headers = ['평가번호', '사업장', '단위공정/활동', '환경측면(Aspect)', '환경영향(Impact)',
               '빈도', '심각도', '중요성점수', '중요성판정', '현재 통제/관리 대책', '평가부서', '평가일자']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h); c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _THIN

    for i, a in enumerate(aspects, 3):
        row = [a.aspect_number, a.site or '제조공장', a.process, a.aspect, a.impact,
               a.frequency, a.severity, a.significance,
               '중요 환경측면' if a.is_significant else '일반 관리',
               a.control_measure or '', deptmap.get(a.department_id, '-'),
               str(a.evaluation_date) if a.evaluation_date else '']
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = _CENTER if col in (1, 2, 6, 7, 8, 9, 12) else _LEFT
            c.border = _THIN
            if a.is_significant and col == 9:
                c.fill = PatternFill('solid', fgColor='fecaca')

    for col, w in enumerate([13, 10, 26, 28, 28, 6, 7, 9, 12, 40, 12, 12], 1):
        ws.column_dimensions[ws.cell(2, col).column_letter].width = w
    ws.freeze_panes = 'A3'
    tag = f'_{site}' if site else ''
    return _xlsx_response(wb, f'환경측면평가{tag}_{datetime.now().strftime("%Y%m%d")}.xlsx')


@hse_bp.route('/risk')
@login_required
def risk():
    site = request.args.get('site', '')
    query = RiskAssessment.query
    if site:
        query = query.filter_by(site=site)
    # id 기준 정렬(고유·불변) — created_at은 대량삽입 시 중복돼 정렬이 흔들릴 수 있음
    risks = query.order_by(RiskAssessment.id.desc()).all()
    risk_ids = [r.id for r in risks]
    # 개선조치 매핑 (위험별 최신) + 현황 집계
    imp_map = {}
    imps = []
    if risk_ids:
        imps = RiskImprovement.query.filter(RiskImprovement.risk_id.in_(risk_ids)) \
            .order_by(RiskImprovement.created_at.desc()).all()
        for imp in imps:
            imp_map.setdefault(imp.risk_id, imp)   # desc 정렬이라 첫 항목=최신
    stats = {
        '상': sum(1 for r in risks if r.risk_grade == '상'),
        '중': sum(1 for r in risks if r.risk_grade == '중'),
        '하': sum(1 for r in risks if r.risk_grade == '하'),
        'improving': sum(1 for i in imps if i.status in ('in_progress', 'completed')),
        'improved': sum(1 for i in imps if i.status == 'approved'),
    }
    return render_template('hse/risk.html', risks=risks, site=site,
                           stats=stats, imp_map=imp_map)


@hse_bp.route('/risk/new', methods=['GET', 'POST'])
@login_required
def new_risk():
    if not current_user.has_edit():
        flash('위험성평가 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.risk'))
    if request.method == 'POST':
        probability = int(request.form.get('probability', 1))
        severity = int(request.form.get('severity', 1))
        risk_level = probability * severity
        if risk_level >= 15:
            risk_grade = '상'
        elif risk_level >= 8:
            risk_grade = '중'
        else:
            risk_grade = '하'

        year = datetime.now().year
        count = RiskAssessment.query.filter(
            RiskAssessment.risk_number.like(f'RISK-{year}-%')
        ).count() + 1

        risk = RiskAssessment(
            risk_number=f'RISK-{year}-{count:03d}',
            site=request.form.get('site', '제조공장'),
            work_area=request.form.get('work_area'),
            work_type=request.form.get('work_type'),
            hazard=request.form.get('hazard'),
            risk_scenario=request.form.get('risk_scenario'),
            probability=probability,
            severity=severity,
            risk_level=risk_level,
            risk_grade=risk_grade,
            control_measure=request.form.get('control_measure'),
            assessment_date=date.today(),
            created_by_id=current_user.id,
        )
        db.session.add(risk)
        db.session.commit()
        flash(f'위험성평가가 등록되었습니다. (위험등급: {risk_grade})', 'success')
        return redirect(url_for('hse.risk'))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('hse/new_risk.html', departments=departments, item=None)


@hse_bp.route('/risk/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def edit_risk(rid):
    risk = RiskAssessment.query.get_or_404(rid)
    if not current_user.has_edit() or not (current_user.is_admin() or risk.created_by_id == current_user.id):
        flash('수정 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('hse.risk'))
    if request.method == 'POST':
        probability = int(request.form.get('probability', 1))
        severity = int(request.form.get('severity', 1))
        risk_level = probability * severity
        risk_grade = '상' if risk_level >= 15 else ('중' if risk_level >= 8 else '하')
        risk.site = request.form.get('site', risk.site)
        risk.work_area = request.form.get('work_area')
        risk.work_type = request.form.get('work_type')
        risk.hazard = request.form.get('hazard')
        risk.risk_scenario = request.form.get('risk_scenario')
        risk.probability = probability
        risk.severity = severity
        risk.risk_level = risk_level
        risk.risk_grade = risk_grade
        risk.control_measure = request.form.get('control_measure')
        db.session.commit()
        flash(f'위험성평가 [{risk.risk_number}]이(가) 수정되었습니다. (위험등급: {risk_grade})', 'success')
        return redirect(url_for('hse.risk'))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('hse/new_risk.html', departments=departments, item=risk)


@hse_bp.route('/risk/<int:rid>/delete', methods=['POST'])
@login_required
def delete_risk(rid):
    from deletion import can_delete
    from models import AuditTrail
    risk = RiskAssessment.query.get_or_404(rid)
    if not can_delete(risk, 'created_by_id'):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('hse.risk'))
    db.session.add(AuditTrail(user_id=current_user.id, action='위험성평가삭제',
        target_type='risk', target_id=risk.id, target_name=risk.risk_number))
    db.session.delete(risk)
    db.session.commit()
    flash(f'위험성평가 [{risk.risk_number}]이(가) 삭제되었습니다.', 'info')
    return redirect(url_for('hse.risk'))


@hse_bp.route('/risk/export')
@login_required
def export_risk():
    site = request.args.get('site', '')
    query = RiskAssessment.query
    if site:
        query = query.filter_by(site=site)
    risks = query.order_by(RiskAssessment.site, RiskAssessment.risk_number).all()
    deptmap = {d.id: d.name for d in Department.query.all()}

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '위험성평가'
    ws.merge_cells('A1:M1')
    ws['A1'] = '주식회사 누리보이스 안전보건 위험성평가 등록부' + (f' ({site})' if site else '')
    ws['A1'].font = _TITLE_FONT; ws['A1'].alignment = _CENTER; ws.row_dimensions[1].height = 30

    headers = ['평가번호', '사업장', '작업구역', '작업구분', '유해·위험요인(Hazard)', '예상 재해 시나리오',
               '가능성', '중대성', '위험성', '등급', '위험 감소대책', '평가부서', '평가일']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h); c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _THIN

    grade_color = {'상': 'fecaca', '중': 'fde68a', '하': 'd1fae5'}
    for i, r in enumerate(risks, 3):
        row = [r.risk_number, r.site or '제조공장', r.work_area, r.work_type, r.hazard, r.risk_scenario,
               r.probability, r.severity, r.risk_level, r.risk_grade,
               r.control_measure or '', deptmap.get(r.department_id, '-'),
               str(r.assessment_date) if r.assessment_date else '']
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = _CENTER if col in (1, 2, 7, 8, 9, 10, 13) else _LEFT
            c.border = _THIN
            if col == 10:
                c.fill = PatternFill('solid', fgColor=grade_color.get(r.risk_grade, 'ffffff'))

    for col, w in enumerate([13, 10, 20, 18, 26, 30, 7, 7, 7, 6, 40, 12, 12], 1):
        ws.column_dimensions[ws.cell(2, col).column_letter].width = w
    ws.freeze_panes = 'A3'
    tag = f'_{site}' if site else ''
    return _xlsx_response(wb, f'위험성평가{tag}_{datetime.now().strftime("%Y%m%d")}.xlsx')


@hse_bp.route('/legal/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_legal(item_id):
    from deletion import can_delete
    from models import AuditTrail
    item = LegalRegister.query.get_or_404(item_id)
    if not can_delete(item, 'created_by_id'):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('hse.legal'))
    db.session.add(AuditTrail(user_id=current_user.id, action='법규삭제',
        target_type='legal', target_id=item.id, target_name=item.law_name))
    db.session.delete(item)   # 조항(clauses) cascade 삭제
    db.session.commit()
    flash('법규 등록부 항목이 삭제되었습니다.', 'info')
    return redirect(url_for('hse.legal'))


@hse_bp.route('/legal/export')
@login_required
def export_legal():
    """법규 등록부 — 법규 목록 + 조·항별 세부 조항까지 포함한 Excel."""
    items = LegalRegister.query.order_by(LegalRegister.reg_number).all()
    deptmap = {d.id: d.name for d in Department.query.all()}
    status_kr = {'compliant': '준수', 'partial': '부분준수', 'non_compliant': '미준수', 'na': '해당없음'}

    wb = openpyxl.Workbook()

    # ── 시트1: 법규 목록 ──
    ws = wb.active; ws.title = '법규 목록'
    ws.merge_cells('A1:J1')
    ws['A1'] = '주식회사 누리보이스 법규 등록부 (ISO 14001/45001)'
    ws['A1'].font = _TITLE_FONT; ws['A1'].alignment = _CENTER; ws.row_dimensions[1].height = 30
    h1 = ['등록번호', '법규명', '유형', 'ISO규격', '적용조항', '적용부서', '준수현황', '준수 비고', '최근검토일', '차기검토일']
    for col, h in enumerate(h1, 1):
        c = ws.cell(2, col, h); c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _THIN
    st_color = {'compliant': 'd1fae5', 'partial': 'fde68a', 'non_compliant': 'fecaca', 'na': 'e5e7eb'}
    for i, it in enumerate(items, 3):
        row = [it.reg_number, it.law_name, it.law_type or '', it.iso_standard or '',
               it.applicable_clause or '', it.applicable_dept or (deptmap.get(it.department_id, '-')),
               status_kr.get(it.compliance_status, it.compliance_status or '-'), it.compliance_note or '',
               str(it.last_review_date) if it.last_review_date else '',
               str(it.next_review_date) if it.next_review_date else '']
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = _CENTER if col in (1, 3, 4, 7, 9, 10) else _LEFT; c.border = _THIN
            if col == 7:
                c.fill = PatternFill('solid', fgColor=st_color.get(it.compliance_status, 'ffffff'))
    for col, w in enumerate([13, 34, 12, 10, 16, 14, 10, 30, 12, 12], 1):
        ws.column_dimensions[ws.cell(2, col).column_letter].width = w
    ws.freeze_panes = 'A3'

    # ── 시트2: 조항 상세 (법규별 조·항) ──
    ws2 = wb.create_sheet('조항 상세')
    ws2.merge_cells('A1:H1')
    ws2['A1'] = '법규별 조·항 세부 내용 (당사 의무사항 포함)'
    ws2['A1'].font = _TITLE_FONT; ws2['A1'].alignment = _CENTER; ws2.row_dimensions[1].height = 30
    h2 = ['등록번호', '법규명', '조', '항', '조항 제목', '조문 내용', '당사 의무사항', '법령원문 링크']
    for col, h in enumerate(h2, 1):
        c = ws2.cell(2, col, h); c.fill = _HDR_FILL; c.font = _HDR_FONT; c.alignment = _CENTER; c.border = _THIN
    r = 3
    for it in items:
        clauses = it.clauses.order_by(LegalClause.sort_order, LegalClause.id).all()
        if not clauses:
            continue
        for cl in clauses:
            row = [it.reg_number, it.law_name, cl.article or '', cl.paragraph or '',
                   cl.clause_title or '', cl.content or '', cl.our_obligation or '', cl.reference_url or '']
            for col, val in enumerate(row, 1):
                c = ws2.cell(r, col, val)
                c.alignment = _CENTER if col in (1, 3, 4) else _LEFT; c.border = _THIN
            r += 1
    if r == 3:
        ws2.merge_cells('A3:H3'); ws2['A3'] = '등록된 세부 조항이 없습니다.'; ws2['A3'].alignment = _CENTER
    for col, w in enumerate([13, 30, 8, 8, 24, 50, 40, 24], 1):
        ws2.column_dimensions[ws2.cell(2, col).column_letter].width = w
    ws2.freeze_panes = 'A3'

    return _xlsx_response(wb, f'법규등록부_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ── 법규 등록부 ────────────────────────────────────────────

@hse_bp.route('/legal')
@login_required
def legal():
    iso = request.args.get('iso', '')
    status = request.args.get('status', '')
    q = request.args.get('q', '')

    query = LegalRegister.query
    if iso:
        query = query.filter_by(iso_standard=iso)
    if status:
        query = query.filter_by(compliance_status=status)
    if q:
        query = query.filter(LegalRegister.law_name.contains(q))

    items = query.order_by(LegalRegister.created_at.desc()).all()
    non_compliant = LegalRegister.query.filter(
        LegalRegister.compliance_status.in_(['non_compliant', 'partial'])
    ).count()
    return render_template('hse/legal.html', items=items, non_compliant=non_compliant,
                           iso=iso, status=status, q=q)


@hse_bp.route('/legal/new', methods=['GET', 'POST'])
@login_required
def new_legal():
    if not current_user.has_edit():
        flash('법규 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.legal'))
    if request.method == 'POST':
        year = datetime.now().year
        count = LegalRegister.query.filter(
            LegalRegister.reg_number.like(f'LAW-{year}-%')
        ).count() + 1
        reg_number = f'LAW-{year}-{count:03d}'

        enacted_str = request.form.get('enacted_date')
        last_review_str = request.form.get('last_review_date')
        next_review_str = request.form.get('next_review_date')

        item = LegalRegister(
            reg_number=reg_number,
            law_name=request.form.get('law_name'),
            law_type=request.form.get('law_type'),
            iso_standard=request.form.get('iso_standard'),
            applicable_clause=request.form.get('applicable_clause'),
            applicable_dept=request.form.get('applicable_dept'),
            compliance_status=request.form.get('compliance_status', 'compliant'),
            compliance_note=request.form.get('compliance_note'),
            enacted_date=datetime.strptime(enacted_str, '%Y-%m-%d').date() if enacted_str else None,
            last_review_date=datetime.strptime(last_review_str, '%Y-%m-%d').date() if last_review_str else date.today(),
            next_review_date=datetime.strptime(next_review_str, '%Y-%m-%d').date() if next_review_str else None,
            department_id=request.form.get('department_id', type=int),
            created_by_id=current_user.id,
        )
        db.session.add(item)
        db.session.commit()
        flash(f'법규 [{reg_number}] {item.law_name}이(가) 등록되었습니다.', 'success')
        return redirect(url_for('hse.legal'))

    departments = Department.query.filter_by(is_active=True).all()
    return render_template('hse/new_legal.html', departments=departments, today=date.today())


@hse_bp.route('/legal/<int:item_id>/update-status', methods=['POST'])
@login_required
def update_legal_status(item_id):
    item = LegalRegister.query.get_or_404(item_id)
    if not current_user.has_legal_edit():
        flash('준수 현황 변경 권한이 없습니다. (내부심사자 또는 관리자만 가능)', 'danger')
        return redirect(url_for('hse.legal'))
    item.compliance_status = request.form.get('compliance_status', item.compliance_status)
    item.compliance_note = request.form.get('compliance_note', item.compliance_note)
    item.last_review_date = date.today()
    db.session.commit()
    flash('준수 현황이 업데이트되었습니다.', 'success')
    return redirect(url_for('hse.legal'))


@hse_bp.route('/legal/<int:item_id>')
@login_required
def legal_detail(item_id):
    """법규 상세 — 조·항별 상세 법령 내용 확인"""
    item = LegalRegister.query.get_or_404(item_id)
    clauses = item.clauses.order_by(LegalClause.sort_order, LegalClause.id).all()
    return render_template('hse/legal_detail.html', item=item, clauses=clauses)


@hse_bp.route('/legal/<int:item_id>/clause/add', methods=['POST'])
@login_required
def add_clause(item_id):
    """조항 상세 추가"""
    item = LegalRegister.query.get_or_404(item_id)
    if not current_user.has_edit():
        flash('조항 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.legal_detail', item_id=item_id))

    max_order = db.session.query(db.func.max(LegalClause.sort_order))\
        .filter_by(legal_id=item_id).scalar() or 0
    clause = LegalClause(
        legal_id=item_id,
        article=request.form.get('article'),
        paragraph=request.form.get('paragraph'),
        clause_title=request.form.get('clause_title'),
        content=request.form.get('content'),
        our_obligation=request.form.get('our_obligation'),
        reference_url=request.form.get('reference_url'),
        sort_order=max_order + 1,
    )
    db.session.add(clause)
    db.session.commit()
    flash(f'{clause.article or "조항"} 상세가 추가되었습니다.', 'success')
    return redirect(url_for('hse.legal_detail', item_id=item_id))


@hse_bp.route('/legal/clause/<int:clause_id>/edit', methods=['POST'])
@login_required
def edit_clause(clause_id):
    """조항 상세 수정"""
    clause = LegalClause.query.get_or_404(clause_id)
    if not current_user.has_edit():
        flash('수정 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.legal_detail', item_id=clause.legal_id))
    clause.article = request.form.get('article')
    clause.paragraph = request.form.get('paragraph')
    clause.clause_title = request.form.get('clause_title')
    clause.content = request.form.get('content')
    clause.our_obligation = request.form.get('our_obligation')
    clause.reference_url = request.form.get('reference_url')
    db.session.commit()
    flash('조항 상세가 수정되었습니다.', 'success')
    return redirect(url_for('hse.legal_detail', item_id=clause.legal_id))


@hse_bp.route('/legal/clause/<int:clause_id>/delete', methods=['POST'])
@login_required
def delete_clause(clause_id):
    """조항 상세 삭제"""
    clause = LegalClause.query.get_or_404(clause_id)
    legal_id = clause.legal_id
    if not current_user.has_edit():
        flash('삭제 권한이 없습니다.', 'danger')
        return redirect(url_for('hse.legal_detail', item_id=legal_id))
    db.session.delete(clause)
    db.session.commit()
    flash('조항 상세가 삭제되었습니다.', 'info')
    return redirect(url_for('hse.legal_detail', item_id=legal_id))
