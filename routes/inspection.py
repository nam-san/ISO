import io, os
from datetime import datetime, date, timedelta

from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user

from models import (db, InspectionType, InspectionRecord,
                    Department, AuditTrail)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

inspection_bp = Blueprint('inspection', __name__)


# ════════════════════════════════════════════════════════════
# 대한민국 일반 제조업 법정 정기점검·측정 기본 주기 시드
#   (명칭, 분류, 근거법령, 주기[개월])
#   ※ 주기는 사업장 규모·유해인자에 따라 달라질 수 있어 관리자가 조정 가능
# ════════════════════════════════════════════════════════════
INSPECTION_SEED = [
    # ── 소방 ──
    ('소방시설 작동점검',      '소방',     '화재의 예방 및 안전관리에 관한 법률', 12),
    ('소방시설 종합점검',      '소방',     '화재의 예방 및 안전관리에 관한 법률', 12),
    ('소방훈련 및 교육',       '소방',     '화재의 예방 및 안전관리에 관한 법률(소방계획)', 12),
    ('피난·방화시설 점검',     '소방',     '화재의 예방 및 안전관리에 관한 법률', 6),
    # ── 안전보건 ──
    ('작업환경측정(정기)',     '안전보건', '산업안전보건법 제125조', 6),
    ('일반건강진단(비사무직)', '안전보건', '산업안전보건법 제129조', 12),
    ('일반건강진단(사무직)',   '안전보건', '산업안전보건법 제129조', 24),
    ('특수건강진단',           '안전보건', '산업안전보건법 제130조(유해인자별 상이)', 12),
    ('위험성평가(정기)',       '안전보건', '산업안전보건법 제36조', 12),
    ('안전검사(위험기계·기구)','안전보건', '산업안전보건법 제93조(프레스·크레인·압력용기 등)', 24),
    ('근로자 정기 안전보건교육','안전보건', '산업안전보건법 제29조(매 분기)', 3),
    # ── 환경 ──
    ('대기오염물질 자가측정',  '환경',     '대기환경보전법 시행규칙 별표11(배출구 종별)', 3),
    ('수질오염물질 자가측정',  '환경',     '물환경보전법 시행규칙(배출규모별)', 3),
    ('소음·진동 측정',         '환경',     '소음·진동관리법 / 작업환경측정 연계', 6),
    ('폐기물 처리실적 보고',   '환경',     '폐기물관리법(연 1회 실적보고)', 12),
    ('온실가스·에너지 명세서', '환경',     '온실가스 목표관리제(해당 사업장)', 12),
    # ── 전기·가스 ──
    ('전기설비 정기검사',      '전기·가스','전기안전관리법(자가용 전기설비)', 36),
    ('가스시설 정기검사',      '전기·가스','고압가스 안전관리법 / 액화석유가스법', 12),
    ('위험물 정기점검',        '전기·가스','위험물안전관리법 제18조', 12),
    # ── 기타 ──
    ('승강기 정기검사',        '기타',     '승강기 안전관리법 제32조', 12),
    ('어린이집·구내식당 위생', '기타',     '식품위생법(구내식당 운영 시)', 6),
]

CATEGORY_ICON = {
    '소방': '🧯', '안전보건': '⛑️', '환경': '🌿', '전기·가스': '⚡', '기타': '📋',
}


def seed_inspection_types():
    """법정점검 항목 마스터 시드 (최초 1회)."""
    if InspectionType.query.count() > 0:
        return
    for i, (name, cat, basis, cycle) in enumerate(INSPECTION_SEED):
        db.session.add(InspectionType(
            name=name, category=cat, legal_basis=basis,
            cycle_months=cycle, lead_alert_days=30,
            is_active=True, sort_order=i,
        ))
    db.session.commit()


def _status_of(itype, today):
    """차기 예정일 기준 상태 판정."""
    if not itype.next_due_date:
        return 'unset', None
    days = (itype.next_due_date - today).days
    if days < 0:
        return 'overdue', days
    if days <= itype.lead_alert_days:
        return 'soon', days
    return 'ok', days


# ── 점검 일정 보드 (메인) ────────────────────────────────────
@inspection_bp.route('/')
@login_required
def index():
    today = date.today()
    cat = request.args.get('category', '')

    query = InspectionType.query.filter_by(is_active=True)
    if cat:
        query = query.filter_by(category=cat)
    types = query.order_by(InspectionType.sort_order).all()

    board = []
    overdue = soon = 0
    for t in types:
        status, days = _status_of(t, today)
        if status == 'overdue': overdue += 1
        elif status == 'soon':  soon += 1
        board.append({'t': t, 'status': status, 'days': days})

    # 정렬: 초과 > 임박 > 정상 > 미설정, 같은 그룹은 예정일 빠른 순
    order = {'overdue': 0, 'soon': 1, 'ok': 2, 'unset': 3}
    board.sort(key=lambda x: (order[x['status']],
                              x['t'].next_due_date or date.max))

    categories = ['소방', '안전보건', '환경', '전기·가스', '기타']
    return render_template('inspection/index.html',
        board=board, today=today, overdue=overdue, soon=soon,
        categories=categories, cat=cat, icons=CATEGORY_ICON)


# ── 점검 실시 기록 등록 ──────────────────────────────────────
@inspection_bp.route('/type/<int:type_id>/perform', methods=['GET', 'POST'])
@login_required
def perform(type_id):
    itype = InspectionType.query.get_or_404(type_id)
    if not current_user.has_edit():
        flash('점검 기록 등록 권한이 없습니다.', 'danger')
        return redirect(url_for('inspection.index'))

    if request.method == 'POST':
        performed = datetime.strptime(request.form['performed_date'], '%Y-%m-%d').date()
        # 차기 예정일: 입력값 우선, 없으면 주기로 자동계산
        nd_str = request.form.get('next_due_date')
        if nd_str:
            next_due = datetime.strptime(nd_str, '%Y-%m-%d').date()
        else:
            next_due = _add_months(performed, itype.cycle_months)

        rec = InspectionRecord(
            inspection_type_id=itype.id,
            performed_date=performed,
            next_due_date=next_due,
            result=request.form.get('result', '적합'),
            performer=request.form.get('performer'),
            department_id=request.form.get('department_id', type=int),
            finding=request.form.get('finding'),
            action=request.form.get('action'),
            created_by_id=current_user.id,
        )

        file = request.files.get('file')
        if file and file.filename:
            from config import Config
            from werkzeug.utils import secure_filename
            from utils import allowed_file
            if not allowed_file(file.filename, Config.ALLOWED_EXTENSIONS):
                flash('허용되지 않는 파일 형식입니다. (PDF/Office/한글/이미지만 가능)', 'danger')
                return redirect(request.url)
            fn = f"INS_{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(file.filename)}"
            fp = os.path.join(Config.UPLOAD_FOLDER, fn)
            file.save(fp)
            rec.file_path = fp
            rec.file_name = file.filename

        db.session.add(rec)

        # 마스터 일정 갱신
        itype.last_performed_date = performed
        itype.next_due_date = next_due

        db.session.add(AuditTrail(user_id=current_user.id, action='법정점검실시',
                                  target_type='inspection', target_id=itype.id,
                                  target_name=itype.name))
        db.session.commit()
        flash(f'[{itype.name}] 점검 기록이 등록되었습니다. 차기 예정일: {next_due}', 'success')
        return redirect(url_for('inspection.index'))

    today = date.today()
    suggested_next = _add_months(today, itype.cycle_months)
    departments = Department.query.filter_by(is_active=True).all()
    return render_template('inspection/perform.html',
        itype=itype, today=today, suggested_next=suggested_next,
        departments=departments)


# ── 점검 이력 ────────────────────────────────────────────────
@inspection_bp.route('/history')
@inspection_bp.route('/history/<int:type_id>')
@login_required
def history(type_id=None):
    query = InspectionRecord.query
    itype = None
    if type_id:
        itype = InspectionType.query.get_or_404(type_id)
        query = query.filter_by(inspection_type_id=type_id)
    records = query.order_by(InspectionRecord.performed_date.desc()).all()
    types = InspectionType.query.order_by(InspectionType.sort_order).all()
    return render_template('inspection/history.html',
        records=records, itype=itype, types=types, today=date.today())


# ── 점검 항목·주기 관리 (관리자) ─────────────────────────────
@inspection_bp.route('/types', methods=['GET', 'POST'])
@login_required
def types():
    if not current_user.is_admin():
        flash('주기 관리는 관리자 권한이 필요합니다.', 'danger')
        return redirect(url_for('inspection.index'))

    if request.method == 'POST':
        action = request.form.get('action_type')

        if action == 'update_cycles':
            # 일괄 주기 변경
            changed = 0
            for t in InspectionType.query.all():
                key = f't_{t.id}_cycle'
                if key in request.form:
                    new_cycle = request.form.get(key, type=int)
                    new_alert = request.form.get(f't_{t.id}_alert', type=int)
                    new_active = request.form.get(f't_{t.id}_active') == 'on'
                    if new_cycle and (t.cycle_months != new_cycle
                                      or t.lead_alert_days != new_alert
                                      or t.is_active != new_active):
                        t.cycle_months = new_cycle
                        t.lead_alert_days = new_alert or 30
                        t.is_active = new_active
                        changed += 1
            db.session.commit()
            flash(f'{changed}개 항목의 주기 설정이 변경되었습니다.', 'success')

        elif action == 'add_type':
            t = InspectionType(
                name=request.form['name'],
                category=request.form.get('category', '기타'),
                legal_basis=request.form.get('legal_basis'),
                cycle_months=request.form.get('cycle_months', type=int) or 12,
                lead_alert_days=request.form.get('lead_alert_days', type=int) or 30,
                is_active=True,
                sort_order=(db.session.query(db.func.max(InspectionType.sort_order)).scalar() or 0) + 1,
            )
            db.session.add(t)
            db.session.commit()
            flash(f'점검 항목 [{t.name}]이 추가되었습니다.', 'success')

        return redirect(url_for('inspection.types'))

    types = InspectionType.query.order_by(InspectionType.sort_order).all()
    return render_template('inspection/types.html',
        types=types, categories=['소방', '안전보건', '환경', '전기·가스', '기타'],
        icons=CATEGORY_ICON)


# ── Excel 출력 (점검 일정표) ─────────────────────────────────
@inspection_bp.route('/export')
@login_required
def export_excel():
    today = date.today()
    types = InspectionType.query.filter_by(is_active=True).order_by(InspectionType.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '법정점검 일정표'

    hdr_fill = PatternFill('solid', fgColor='1a3a5c')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:H1')
    ws['A1'] = f'법정 정기점검·측정 일정표 (기준일: {today})'
    ws['A1'].font = Font(bold=True, size=14, color='1a3a5c')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ['분류', '점검 항목', '근거 법령', '주기(개월)', '최근 실시일', '차기 예정일', '잔여(일)', '상태']
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin

    status_kr = {'overdue': '기한초과', 'soon': '도래임박', 'ok': '정상', 'unset': '미실시'}
    status_color = {'overdue': 'fecaca', 'soon': 'fef9c3', 'ok': 'd1fae5', 'unset': 'e5e7eb'}

    for i, t in enumerate(types, 3):
        status, days = _status_of(t, today)
        row = [t.category, t.name, t.legal_basis, t.cycle_months,
               str(t.last_performed_date) if t.last_performed_date else '-',
               str(t.next_due_date) if t.next_due_date else '-',
               days if days is not None else '-', status_kr[status]]
        for col, val in enumerate(row, 1):
            c = ws.cell(i, col, val)
            c.alignment = center; c.border = thin
            if col == 8:
                c.fill = PatternFill('solid', fgColor=status_color[status])

    widths = [12, 26, 36, 10, 13, 13, 9, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(2, i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output,
        download_name=f'법정점검일정표_{today.strftime("%Y%m%d")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@inspection_bp.route('/record/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_record(record_id):
    from deletion import can_delete, remove_file
    from models import AuditTrail
    rec = InspectionRecord.query.get_or_404(record_id)
    if not can_delete(rec, 'created_by_id'):
        flash('삭제 권한이 없습니다. (작성자 또는 관리자만)', 'danger')
        return redirect(url_for('inspection.history'))
    remove_file(rec.file_path)
    db.session.add(AuditTrail(user_id=current_user.id, action='법정점검이력삭제',
        target_type='inspection', target_id=rec.id,
        target_name=rec.inspection_type.name if rec.inspection_type else str(rec.id)))
    db.session.delete(rec)
    db.session.commit()
    flash('법정점검 실시 이력이 삭제되었습니다.', 'info')
    return redirect(url_for('inspection.history'))


# ── 개월 더하기 유틸 ──────────────────────────────────────────
def _add_months(d, months):
    month = d.month - 1 + months
    year = d.year + month // 12
    month = month % 12 + 1
    import calendar
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


# ── 알림 조회 (대시보드/시스템 연계용) ───────────────────────
def get_inspection_alerts():
    """도래 임박 + 기한 초과 점검 항목 반환."""
    today = date.today()
    types = InspectionType.query.filter_by(is_active=True).all()
    overdue, soon = [], []
    for t in types:
        status, days = _status_of(t, today)
        if status == 'overdue':
            overdue.append((t, days))
        elif status == 'soon':
            soon.append((t, days))
    overdue.sort(key=lambda x: x[1])
    soon.sort(key=lambda x: x[1])
    return overdue, soon
